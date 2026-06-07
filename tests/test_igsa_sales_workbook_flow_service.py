from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlmodel import SQLModel, Session, create_engine, select

import app.services.igsa_sales_workbook_flow_service as igsa_workbook_service_module
from app.models import Cliente, Distribuidor, IngredienteIreks, ReferenciaDistribuidor, VentaImportLote, VentaMensualRaw
from app.services.igsa_sales_workbook_flow_service import IgsaSalesWorkbookFlowService


@pytest.fixture()
def isolated_engine(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'igsa-workbook.db'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(engine)
    monkeypatch.setattr(igsa_workbook_service_module, "engine", engine)
    return engine


def _seed_preview_data(session: Session) -> tuple[str, str, str]:
    distributor_id = "dist-1"
    client_id = "cli-1"
    article_id = "art-1"
    session.add(
        Distribuidor(
            distribuidor_id=distributor_id,
            distribuidor_codigo=1,
            distribuidor_razon_social="IGSA Canarias",
            distribuidor_nombre_comercial="IGSA",
        )
    )
    session.add(Cliente(cliente_id=client_id, cliente_codigo=1, cliente_nombre_comercial="Cliente", distribuidor_id=distributor_id))
    session.add(
        IngredienteIreks(
            articulo_id=article_id,
            almacen_id="alm-1",
            distribuidor_id=distributor_id,
            articulo_referencia="D123",
            articulo_referencia_corta="D123",
            articulo_descripcion="Producto IREKS",
            articulo_envase_peso_total=2.5,
            articulo_envase_peso=2.5,
        )
    )
    session.add(
        ReferenciaDistribuidor(
            articulo_id=article_id,
            distribuidor_id=distributor_id,
            articulo_referencia_distribuidor="D123",
            articulo_descripcion_distribuidor="Producto IREKS",
        )
    )
    session.commit()
    return client_id, distributor_id, article_id


def _build_workbook(path: Path) -> None:
    workbook = Workbook()
    ws1 = workbook.active
    ws1.title = "Venta"
    ws1["A1"] = "Mayo 2026"
    ws1["A2"] = "D123"
    ws1["B2"] = "D123"
    ws1["C2"] = "Producto IREKS"
    ws1["D2"] = 2.5
    ws1["E2"] = 4
    ws1["F2"] = "L.01"

    ws2 = workbook.create_sheet("Promo")
    ws2["A1"] = "Mayo 2026"
    ws2["A2"] = "D123"
    ws2["B2"] = "D123"
    ws2["C2"] = "Producto IREKS promo"
    ws2["D2"] = 2.5
    ws2["E2"] = 2
    ws2["F2"] = "2 L.02"

    workbook.save(path)


def test_parse_workbook_by_sheets_and_preview_and_import(isolated_engine) -> None:
    service = IgsaSalesWorkbookFlowService()
    workbook_path = Path(isolated_engine.url.database).with_name("igsa.xlsx")
    _build_workbook(workbook_path)

    rows, errors = service.parse_igsa_workbook_by_sheets(workbook_path)
    assert errors == []
    assert len(rows) == 2
    assert rows[0].periodo == "2026-05"
    assert rows[0].tipo == "venta"
    assert rows[1].tipo == "promocion"
    assert rows[0].cantidad_lote == 4.0

    with Session(isolated_engine) as session:
        client_id, _distributor_id, _article_id = _seed_preview_data(session)

    preview_rows, preview_errors = service.build_igsa_workbook_preview(rows, client_id)
    assert preview_errors == []
    assert len(preview_rows) == 2
    assert preview_rows[0]["articulo_id"] == "art-1"
    assert preview_rows[0]["tot_kg"] == pytest.approx(10.0)

    sync_calls: list[int] = []
    result = service.import_igsa_workbook_lines(
        rows,
        client_id,
        sync_warehouse_callback=lambda _session, inserted_rows: sync_calls.append(len(inserted_rows)),
    )
    assert result.ok is True
    assert result.imported == 2
    assert result.incidencias == 0
    assert "Importacion IGSA (libro por hojas) completada." in result.message
    assert sync_calls == [2]

    with Session(isolated_engine) as session:
        assert session.exec(select(VentaImportLote).where(VentaImportLote.fuente == "igsa_book")).first() is not None
        raw_rows = list(session.exec(select(VentaMensualRaw).where(VentaMensualRaw.fuente == "igsa_book")))
        assert len(raw_rows) == 2

    second = service.import_igsa_workbook_lines(rows, client_id)
    assert second.ok is True
    assert second.imported == 0
    assert second.message == "El libro ya estaba importado."


def test_parse_workbook_reports_missing_period_and_missing_file(tmp_path: Path) -> None:
    service = IgsaSalesWorkbookFlowService()

    missing_rows, missing_errors = service.parse_igsa_workbook_by_sheets(tmp_path / "missing.xlsx")
    assert missing_rows == []
    assert missing_errors and "No existe el archivo" in missing_errors[0]

    workbook_path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)

    rows, errors = service.parse_igsa_workbook_by_sheets(workbook_path)
    assert rows == []
    assert errors and "no se pudo inferir periodo" in errors[0].lower()


def test_import_workbook_lines_validates_inputs(isolated_engine) -> None:
    service = IgsaSalesWorkbookFlowService()
    result = service.import_igsa_workbook_lines([], "")
    assert result.ok is False
    assert "Cliente/Distribuidor IGSA no valido." == result.message
