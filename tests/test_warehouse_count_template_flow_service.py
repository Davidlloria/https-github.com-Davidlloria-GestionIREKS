from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.warehouse_count_template_flow_service import WarehouseCountTemplateFlowService


def test_prepare_export_rows_builds_export_ready_data() -> None:
    service = WarehouseCountTemplateFlowService()

    result = service.prepare_export_rows(
        [
            {
                "articulo_id": "A1",
                "ref": "REF-1",
                "nombre": "Producto 1",
                "lote": "L1",
                "caducidad": "2026-01-15",
                "teorico_uds": 12.5,
                "conteo_uds": "",
            }
        ]
    )

    assert result.status == "ready"
    assert result.export_rows == [
        {
            "articulo_id": "A1",
            "ref": "REF-1",
            "nombre": "Producto 1",
            "lote": "L1",
            "caducidad": "2026-01-15",
            "teorico_uds": "12.5",
            "conteo_uds": "",
        }
    ]


def test_prepare_export_rows_without_data_marks_empty() -> None:
    service = WarehouseCountTemplateFlowService()

    result = service.prepare_export_rows([])

    assert result.status == "empty"
    assert result.export_rows == []


def test_read_import_file_returns_ready_mapping(tmp_path: Path) -> None:
    service = WarehouseCountTemplateFlowService()
    file_path = tmp_path / "conteo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo"
    ws.append(["articulo_id", "ref", "nombre", "lote", "caducidad", "teorico_uds", "conteo_uds"])
    ws.append(["A1", "REF-1", "Producto 1", "L1", "2026-01-15", 12, 9])
    ws.append(["A2", "REF-2", "Producto 2", None, None, 5, "7"])
    wb.save(file_path)

    result = service.read_import_file(file_path)

    assert result.status == "ready"
    assert result.imported_count == 2
    assert result.mapping == {("A1", "L1"): "9", ("A2", ""): "7"}


def test_read_import_file_missing_columns_returns_error(tmp_path: Path) -> None:
    service = WarehouseCountTemplateFlowService()
    file_path = tmp_path / "conteo_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo"
    ws.append(["articulo_id", "ref", "nombre"])
    wb.save(file_path)

    result = service.read_import_file(file_path)

    assert result.status == "missing_columns"
    assert "articulo_id" in result.message


def test_read_import_file_empty_file_returns_empty(tmp_path: Path) -> None:
    service = WarehouseCountTemplateFlowService()
    file_path = tmp_path / "conteo_empty.xlsx"
    wb = Workbook()
    wb.save(file_path)

    result = service.read_import_file(file_path)

    assert result.status == "empty"
    assert "no contiene datos" in result.message.lower()


def test_count_template_mapping_skips_blank_count_and_normalizes_lote() -> None:
    service = WarehouseCountTemplateFlowService()

    mapping = service.count_template_mapping(
        [
            ("A1", None, "10"),
            ("A2", "", ""),
            ("", "L3", "5"),
        ],
        0,
        1,
        2,
    )

    assert mapping == {("A1", ""): "10"}


def test_count_template_column_indexes_detects_required_columns() -> None:
    service = WarehouseCountTemplateFlowService()

    assert service.count_template_column_indexes(["articulo_id", "lote", "conteo_uds"]) == (0, 1, 2)
