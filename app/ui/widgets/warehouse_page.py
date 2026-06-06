from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, cast as tcast
from uuid import uuid4

from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QDoubleSpinBox,
    QFormLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)
from openpyxl import Workbook, load_workbook

from app.models import (
    AlmacenMovimiento,
    Cliente,
    Envase,
    Fabricante,
    Familia,
    Subfamilia,
)
from app.services.monthly_orders_service import MonthlyOrdersService
from app.services.warehouse_inventory_adjustment_preparation_service import (
    WarehouseInventoryAdjustmentPreparationService,
)
from app.services.warehouse_manual_move_flow_service import WarehouseManualMoveFlowService
from app.services.warehouse_catalog_service import WarehouseCatalogService
from app.services.warehouse_inventory_service import WarehouseInventoryService
from app.services.warehouse_movement_service import WarehouseMovementService
from app.services.warehouse_reference_service import OtrasRefRow, WarehouseReferenceService
from app.services.warehouse_settings_service import WarehouseSettingsService
from app.ui.widgets.entity_dialog import EntityDialog
from app.ui.widgets.entity_page import EntityPage
from app.ui.widgets.ingredients_page import IngredientsIreksPage
from app.viewmodels.ingredient_viewmodel import IngredientWarehouseViewModel


def _col(expr: object) -> Any:
    return tcast(Any, expr)


class SortableTableWidgetItem(QTableWidgetItem):
    def __init__(self, text: str, sort_value: Any = None) -> None:
        super().__init__(text)
        self.setData(Qt.ItemDataRole.UserRole + 10, sort_value)

    def __lt__(self, other: QTableWidgetItem) -> bool:
        left = self.data(Qt.ItemDataRole.UserRole + 10)
        right = other.data(Qt.ItemDataRole.UserRole + 10) if other is not None else None
        if isinstance(left, (int, float)) and isinstance(right, (int, float)):
            return float(left) < float(right)
        return super().__lt__(other)


def _count_template_column_indexes(header: list[str]) -> tuple[int, int, int]:
    idx_id = header.index("articulo_id") if "articulo_id" in header else -1
    idx_lote = header.index("lote") if "lote" in header else -1
    idx_conteo = header.index("conteo_uds") if "conteo_uds" in header else -1
    return idx_id, idx_lote, idx_conteo


def _count_template_mapping(
    rows: list[tuple[object, ...]],
    idx_id: int,
    idx_lote: int,
    idx_conteo: int,
) -> dict[tuple[str, str], str]:
    mapping: dict[tuple[str, str], str] = {}
    for row in rows:
        art_id = str(row[idx_id] or "").strip()
        lote = str(row[idx_lote] or "").strip()
        conteo_raw = str(row[idx_conteo] or "").strip()
        if not art_id or conteo_raw == "":
            continue
        mapping[(art_id, lote)] = conteo_raw
    return mapping


@dataclass(frozen=True)
class InventoryTabView:
    intro_text: str = "Conteo físico por producto/lote. Edita 'Conteo Uds' y pulsa Aplicar ajustes."
    export_template_button_label: str = "Exportar plantilla"
    import_count_button_label: str = "Importar conteo"
    refresh_button_label: str = "Refrescar"
    prepare_adjustments_button_label: str = "Preparar ajustes"
    apply_adjustments_button_label: str = "Aprobar y aplicar"
    search_label: str = "Buscar"
    search_placeholder: str = "Producto, ref o lote..."
    counter_label: str = "Contador"
    counter_placeholder: str = "Usuario que realiza conteo"
    approver_label: str = "Aprobador"
    approver_placeholder: str = "Usuario que aprueba ajuste"
    pending_label: str = "Pendientes: 0"
    history_title: str = "Historial de inventarios"
    export_history_button_label: str = "Exportar historial"
    log_placeholder: str = "Registro de acciones de mantenimiento..."


class OtrasReferenciasTab(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.reference_service = WarehouseReferenceService()
        self.rows: list[OtrasRefRow] = []
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        header = QLabel("Otras referencias")
        header.setProperty("role", "pageTitle")
        layout.addWidget(header)

        toolbar = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar...")
        self.search_input.textChanged.connect(self.reload)
        toolbar.addWidget(self.search_input, 2)

        self.distributor_filter = QComboBox()
        self.distributor_filter.setMinimumWidth(260)
        self.distributor_filter.currentIndexChanged.connect(self.reload)
        toolbar.addWidget(self.distributor_filter, 1)

        new_btn = QPushButton("Nuevo")
        new_btn.setProperty("btnRole", "success")
        edit_btn = QPushButton("Editar")
        edit_btn.setProperty("btnRole", "warning")
        del_btn = QPushButton("Eliminar")
        del_btn.setProperty("btnRole", "danger")
        import_btn = QPushButton("Importar Excel/CSV")
        import_btn.setProperty("btnRole", "secondary")
        refresh_btn = QPushButton("Refrescar")
        refresh_btn.setProperty("btnRole", "secondary")

        new_btn.clicked.connect(self._new_entity)
        edit_btn.clicked.connect(self._edit_entity)
        del_btn.clicked.connect(self._delete_entity)
        import_btn.clicked.connect(self._import_entities)
        refresh_btn.clicked.connect(self.reload)

        toolbar.addWidget(new_btn)
        toolbar.addWidget(edit_btn)
        toolbar.addWidget(del_btn)
        toolbar.addWidget(import_btn)
        toolbar.addWidget(refresh_btn)
        layout.addLayout(toolbar)

        self.table = QTableWidget(0, 4)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setStyleSheet(
            """
            QTableWidget::item:focus { border: none; outline: 0; }
            QTableWidget::item:selected { color: #0F172A; }
            """
        )
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setHorizontalHeaderLabels(
            ["Ref. fabricante", "Descripcion fab.", "Ref. distribuidor", "Descripcion dist."]
        )
        layout.addWidget(self.table, 1)

    def _schema(self) -> list[dict[str, Any]]:
        return [
            {"name": "articulo_id", "label": "Articulo_ID"},
            {"name": "distribuidor_id", "label": "Distribuidor_ID"},
            {"name": "articulo_referencia_distribuidor", "label": "Articulo_Referencia_Distribuidor"},
            {"name": "articulo_descripcion_distribuidor", "label": "Articulo_Descripcion_Distribuidor"},
        ]

    def _selected_distributor_id(self) -> str:
        return str(self.distributor_filter.currentData() or "")

    def _reload_distributors(self) -> None:
        current = self._selected_distributor_id()
        rows = self.reference_service.list_distributors()
        self.distributor_filter.blockSignals(True)
        self.distributor_filter.clear()
        self.distributor_filter.addItem("Distribuidor (todos)", "")
        for row in rows:
            distribuidor_id = str(row.distribuidor_id or "").strip()
            if not distribuidor_id:
                continue
            label = str(row.distribuidor_nombre_comercial or "").strip() or str(row.distribuidor_razon_social or "").strip()
            if not label:
                label = distribuidor_id
            self.distributor_filter.addItem(label, distribuidor_id)
        idx = self.distributor_filter.findData(current)
        self.distributor_filter.setCurrentIndex(idx if idx >= 0 else 0)
        self.distributor_filter.blockSignals(False)

    def reload(self) -> None:
        self._reload_distributors()
        term = self.search_input.text().strip()
        selected_distributor = self._selected_distributor_id().strip()
        self.rows = self.reference_service.list_references(term=term, distributor_id=selected_distributor)
        self.table.setRowCount(len(self.rows))
        for row_idx, item in enumerate(self.rows):
            values = [
                item.articulo_referencia_fabricante,
                item.articulo_descripcion_fabricante,
                item.articulo_referencia_distribuidor,
                item.articulo_descripcion_distribuidor,
            ]
            for col_idx, value in enumerate(values):
                t_item = QTableWidgetItem(value)
                if col_idx == 0:
                    t_item.setData(Qt.ItemDataRole.UserRole, (item.articulo_id, item.distribuidor_id))
                self.table.setItem(row_idx, col_idx, t_item)

    def _selected_row(self) -> OtrasRefRow | None:
        selected = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected:
            return None
        return self.rows[selected[0].row()]

    def _upsert_reference(self, payload: dict[str, Any], old_key: tuple[str, str] | None = None) -> None:
        self.reference_service.upsert_reference(payload, old_key=old_key)

    def _new_entity(self) -> None:
        dialog = EntityDialog("Nueva: Otras referencias", self._schema(), parent=self)
        if dialog.exec():
            try:
                self._upsert_reference(dialog.get_payload())
                self.reload()
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "Otras referencias", str(exc))

    def _edit_entity(self) -> None:
        row = self._selected_row()
        if not row:
            QMessageBox.warning(self, "Atencion", "Selecciona un registro.")
            return
        initial = {
            "articulo_id": row.articulo_id,
            "distribuidor_id": row.distribuidor_id,
            "articulo_referencia_distribuidor": row.articulo_referencia_distribuidor,
            "articulo_descripcion_distribuidor": row.articulo_descripcion_distribuidor,
        }
        dialog = EntityDialog("Editar: Otras referencias", self._schema(), initial=initial, parent=self)
        if dialog.exec():
            try:
                self._upsert_reference(dialog.get_payload(), old_key=(row.articulo_id, row.distribuidor_id))
                self.reload()
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "Otras referencias", str(exc))

    def _delete_entity(self) -> None:
        row = self._selected_row()
        if not row:
            QMessageBox.warning(self, "Atencion", "Selecciona un registro.")
            return
        answer = QMessageBox.question(self, "Confirmar", "Eliminar referencia seleccionada?")
        if answer != QMessageBox.StandardButton.Yes:
            return
        self.reference_service.delete_reference(row.articulo_id, row.distribuidor_id)
        self.reload()

    def _import_entities(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo",
            "",
            "Archivos de datos (*.xlsx *.xlsm *.csv)",
        )
        if not file_path:
            return
        imported, errors = self.reference_service.import_references(file_path, self._schema())
        self.reload()
        if errors:
            preview = "\n".join(errors[:8])
            extra = "" if len(errors) <= 8 else f"\n... y {len(errors) - 8} errores mas."
            QMessageBox.warning(
                self,
                "Importacion completada con incidencias",
                f"Registros importados: {imported}\nErrores: {len(errors)}\n\n{preview}{extra}",
            )
            return
        QMessageBox.information(self, "Importacion completada", f"Registros importados: {imported}")


class CaducidadTab(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._almacen_id = ""
        self.movement_service = WarehouseMovementService()
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        filters = QHBoxLayout()
        filters.addWidget(QLabel("Caduca desde"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd/MM/yyyy")
        self.date_from.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_from.dateChanged.connect(lambda _d: self.reload())
        filters.addWidget(self.date_from)

        filters.addWidget(QLabel("Caduca hasta"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd/MM/yyyy")
        self.date_to.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.date_to.setDate(QDate(2100, 12, 31))
        self.date_to.dateChanged.connect(lambda _d: self.reload())
        filters.addWidget(self.date_to)

        filters.addWidget(QLabel("Próxima caducidad"))
        self.near_days_combo = QComboBox()
        for days in (7, 15, 30, 45, 60, 90, 120):
            self.near_days_combo.addItem(f"{days} días", days)
        self.near_days_combo.setCurrentIndex(self.near_days_combo.findData(30))
        self.near_days_combo.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.near_days_combo)

        self.mode_combo = QComboBox()
        self.mode_combo.addItem("Caducados + próximos", "both")
        self.mode_combo.addItem("Solo caducados", "expired")
        self.mode_combo.addItem("Solo próximos", "soon")
        self.mode_combo.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.mode_combo)

        clear_btn = QPushButton("Todo")
        clear_btn.setProperty("btnRole", "secondary")
        clear_btn.clicked.connect(self._reset_filters)
        filters.addWidget(clear_btn)
        filters.addStretch(1)
        layout.addLayout(filters)

        self.table = QTableWidget(0, 9)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setStyleSheet("QTableWidget::item:focus { border: none; outline: 0; }")
        header = self.table.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Fixed)
        self.table.setHorizontalHeaderLabels(["Pedido Nº", "Albarán", "Fecha", "Ref.", "Nombre", "Uds", "Kg", "Lote", "Caduca"])
        self.table.setColumnWidth(0, 130)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 95)
        self.table.setColumnWidth(5, 85)
        self.table.setColumnWidth(6, 105)
        self.table.setColumnWidth(7, 105)
        self.table.setColumnWidth(8, 110)
        layout.addWidget(self.table, 1)

    def _reset_filters(self) -> None:
        self.date_from.blockSignals(True)
        self.date_to.blockSignals(True)
        self.near_days_combo.blockSignals(True)
        self.mode_combo.blockSignals(True)
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate(2100, 12, 31))
        self.near_days_combo.setCurrentIndex(self.near_days_combo.findData(30))
        self.mode_combo.setCurrentIndex(self.mode_combo.findData("both"))
        self.date_from.blockSignals(False)
        self.date_to.blockSignals(False)
        self.near_days_combo.blockSignals(False)
        self.mode_combo.blockSignals(False)
        self.reload()

    def set_almacen_filter(self, almacen_id: str) -> None:
        self._almacen_id = str(almacen_id or "").strip()
        self.reload()

    def reload(self) -> None:
        moves, items = self.movement_service.expiration_payload(self._almacen_id)
        peso_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): float(getattr(x, "articulo_envase_peso_total", 0.0) or 0.0) for x in items}
        ref_by_articulo = {
            str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip()
            for x in items
        }
        nombre_by_articulo = {
            str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip()
            for x in items
        }

        from_date = tcast(date, self.date_from.date().toPython())
        to_date = tcast(date, self.date_to.date().toPython())
        if from_date > to_date:
            from_date, to_date = to_date, from_date
        near_days = int(self.near_days_combo.currentData() or 30)
        mode = str(self.mode_combo.currentData() or "both")
        today = date.today()
        alert_limit = today + timedelta(days=near_days)

        rows: list[tuple[AlmacenMovimiento, float, bool, bool]] = []
        # Agrupa por articulo+lote+caducidad para calcular stock neto del lote.
        lot_groups: dict[tuple[str, str, date | None], list[AlmacenMovimiento]] = {}
        for mov in moves:
            key = (
                str(getattr(mov, "articulo_id", "") or "").strip(),
                str(getattr(mov, "articulo_lote", "") or "").strip(),
                getattr(mov, "articulo_caducidad", None),
            )
            lot_groups.setdefault(key, []).append(mov)

        collapsed_moves: list[AlmacenMovimiento] = []
        for _key, lot_moves in lot_groups.items():
            net_qty = sum(float(getattr(x, "cantidad", 0.0) or 0.0) for x in lot_moves)
            # Si el lote ya salió por completo (o quedó en cero), no debe aparecer.
            if net_qty <= 0:
                continue
            # Conserva el movimiento más reciente solo para mostrar metadatos (pedido/albarán/fecha).
            selected = sorted(
                lot_moves,
                key=lambda x: (
                    getattr(x, "fecha_pedido", date.min) or date.min,
                    int(getattr(x, "id", 0) or 0),
                ),
                reverse=True,
            )[0]
            # Sobrescribe cantidad con stock neto vigente del lote.
            setattr(selected, "cantidad", net_qty)
            collapsed_moves.append(selected)

        for mov in collapsed_moves:
            cad = getattr(mov, "articulo_caducidad", None)
            if cad is None or cad < from_date or cad > to_date:
                continue
            is_expired = cad < today
            is_soon = today <= cad <= alert_limit
            if mode == "expired" and not is_expired:
                continue
            if mode == "soon" and not is_soon:
                continue
            if mode == "both" and not (is_expired or is_soon):
                continue
            peso_total = peso_by_articulo.get(str(getattr(mov, "articulo_id", "") or "").strip(), 0.0)
            rows.append((mov, peso_total, is_expired, is_soon))

        self.table.setRowCount(len(rows))
        for row_idx, (mov, peso_total, is_expired, is_soon) in enumerate(rows):
            fecha = mov.fecha_pedido.strftime("%d/%m/%Y") if mov.fecha_pedido else ""
            cantidad = float(getattr(mov, "cantidad", 0.0) or 0.0)
            kg = cantidad * float(peso_total or 0.0)
            caduca = mov.articulo_caducidad.strftime("%d/%m/%Y") if mov.articulo_caducidad else ""
            articulo_id = str(getattr(mov, "articulo_id", "") or "").strip()
            ref = ref_by_articulo.get(articulo_id, "") or articulo_id
            nombre = nombre_by_articulo.get(articulo_id, "") or articulo_id
            values = [
                str(getattr(mov, "pedido_numero", "") or "").strip(),
                str(getattr(mov, "pedido_albaran_numero", "") or "").strip(),
                fecha,
                ref,
                nombre,
                f"{cantidad:.2f}",
                f"{kg:.2f} kg",
                str(getattr(mov, "articulo_lote", "") or "").strip(),
                caduca,
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx in (5, 6):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if is_expired:
                    item.setForeground(QBrush(QColor("#c62828")))
                elif is_soon:
                    item.setForeground(QBrush(QColor("#ef6c00")))
                self.table.setItem(row_idx, col_idx, item)


class MovimientosTab(QWidget):
    def __init__(self, mode: str = "all") -> None:
        super().__init__()
        self._almacen_id = ""
        self._mode = mode  # "all" | "in" | "out"
        self._building_filters = False
        self._movement_by_id: dict[int, AlmacenMovimiento] = {}
        self.movement_service = WarehouseMovementService()
        self.manual_move_flow_service = WarehouseManualMoveFlowService(movement_service=self.movement_service)
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        if self._mode in {"in", "out", "all"}:
            filters = QHBoxLayout()
            filters.addWidget(QLabel("Año"))
            self.year_filter = QComboBox()
            self.year_filter.currentIndexChanged.connect(self.reload)
            filters.addWidget(self.year_filter)

            filters.addWidget(QLabel("Mes"))
            self.month_filter = QComboBox()
            self.month_filter.currentIndexChanged.connect(self.reload)
            filters.addWidget(self.month_filter)

            filters.addWidget(QLabel("Fabricante"))
            self.manufacturer_filter = QComboBox()
            self.manufacturer_filter.currentIndexChanged.connect(self.reload)
            filters.addWidget(self.manufacturer_filter)

            filters.addWidget(QLabel("Familia"))
            self.family_filter = QComboBox()
            self.family_filter.currentIndexChanged.connect(self.reload)
            filters.addWidget(self.family_filter)

            filters.addWidget(QLabel("Subfamilia"))
            self.subfamily_filter = QComboBox()
            self.subfamily_filter.currentIndexChanged.connect(self.reload)
            filters.addWidget(self.subfamily_filter)

            self.occurrence_filter = QLineEdit()
            self.occurrence_filter.setPlaceholderText("Nombre o ref...")
            self.occurrence_filter.textChanged.connect(self.reload)
            if self._mode == "all":
                filters.addWidget(QLabel("Producto"))
                filters.addWidget(self.occurrence_filter)
            filters.addStretch(1)
            layout.addLayout(filters)
        if self._mode in {"in", "out"}:
            actions = QHBoxLayout()
            field_height = self.occurrence_filter.sizeHint().height()
            actions.addWidget(QLabel("Producto"))
            self.occurrence_filter.setMinimumWidth(420)
            actions.addWidget(self.occurrence_filter)
            add_btn = QPushButton("Nueva manual")
            add_btn.setFixedHeight(field_height)
            add_btn.setStyleSheet(
                "QPushButton { background-color: #198754; color: #FFFFFF; border: 1px solid #198754; border-radius: 4px; padding: 0 12px; font-weight: 600; }"
                "QPushButton:hover { background-color: #157347; border-color: #157347; }"
                "QPushButton:pressed { background-color: #125f3b; border-color: #125f3b; }"
            )
            add_btn.clicked.connect(self._create_manual_move)
            edit_btn = QPushButton("Editar manual")
            edit_btn.setFixedHeight(field_height)
            edit_btn.setStyleSheet(
                "QPushButton { background-color: #F59E0B; color: #111827; border: 1px solid #D97706; border-radius: 4px; padding: 0 12px; font-weight: 600; }"
                "QPushButton:hover { background-color: #D97706; border-color: #B45309; color: #FFFFFF; }"
                "QPushButton:pressed { background-color: #B45309; border-color: #92400E; color: #FFFFFF; }"
            )
            edit_btn.clicked.connect(self._edit_manual_move)
            reverse_btn = QPushButton("Anular manual")
            reverse_btn.setFixedHeight(field_height)
            reverse_btn.setStyleSheet(
                "QPushButton { background-color: #DC2626; color: #FFFFFF; border: 1px solid #DC2626; border-radius: 4px; padding: 0 12px; font-weight: 600; }"
                "QPushButton:hover { background-color: #B91C1C; border-color: #B91C1C; }"
                "QPushButton:pressed { background-color: #991B1B; border-color: #991B1B; }"
            )
            reverse_btn.clicked.connect(self._reverse_manual_move)
            actions.addWidget(add_btn)
            actions.addWidget(edit_btn)
            actions.addWidget(reverse_btn)
            actions.addStretch(1)
            layout.addLayout(actions)

        col_count = 7 if self._mode in {"in", "out"} else 8
        self.table = QTableWidget(0, col_count)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setStyleSheet("QTableWidget::item:focus { border: none; outline: 0; }")
        header = self.table.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)

        if self._mode in {"in", "out"}:
            self.table.setHorizontalHeaderLabels(["Fecha", "Ref.", "Nombre", "Uds", "Kg", "Lote", "Concepto"])
            self.table.setColumnWidth(0, 110)
            self.table.setColumnWidth(1, 110)
            self.table.setColumnWidth(2, 520)
            self.table.setColumnWidth(3, 70)
            self.table.setColumnWidth(4, 88)
            self.table.setColumnWidth(5, 105)
            self.table.setColumnWidth(6, 170)
        elif self._mode == "all":
            header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
            self.table.setHorizontalHeaderLabels(["Mes", "Ref.", "Nombre", "Entradas", "Salidas", "Kg Entradas", "Kg Salidas", "Lote"])
            self.table.setColumnWidth(0, 90)
            self.table.setColumnWidth(1, 105)
            self.table.setColumnWidth(2, 430)
            self.table.setColumnWidth(3, 75)
            self.table.setColumnWidth(4, 75)
            self.table.setColumnWidth(5, 95)
            self.table.setColumnWidth(6, 95)
            self.table.setColumnWidth(7, 110)
        else:
            header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
            self.table.setHorizontalHeaderLabels(["Pedido N?", "Albar?n", "Fecha", "Ref.", "Nombre", "Uds", "Kg", "Tipo"])
            self.table.setColumnWidth(0, 130)
            self.table.setColumnWidth(1, 150)
            self.table.setColumnWidth(2, 110)
            self.table.setColumnWidth(3, 95)
            self.table.setColumnWidth(5, 85)
            self.table.setColumnWidth(6, 105)
            self.table.setColumnWidth(7, 100)

        layout.addWidget(self.table, 1)

    def set_almacen_filter(self, almacen_id: str) -> None:
        self._almacen_id = str(almacen_id or "").strip()
        self.reload()

    def _current_filter_data(self, combo: QComboBox, default: str = "") -> str:
        return str(combo.currentData() or default).strip()

    def _concept_from_albaran(self, value: str) -> str:
        text = str(value or "").strip()
        if text.upper().startswith("IGSA"):
            text = text[4:].lstrip(" -_:")
        return text.strip()

    def _reload_out_filters(
        self,
        years: list[int],
        manufacturers: list[tuple[str, str]],
        families: list[tuple[str, str]],
        subfamilies: list[tuple[str, str]],
    ) -> None:
        self._building_filters = True
        try:
            current_year = self._current_filter_data(self.year_filter)
            current_month = self._current_filter_data(self.month_filter, "0")
            current_mfg = self._current_filter_data(self.manufacturer_filter)
            current_fam = self._current_filter_data(self.family_filter)
            current_sub = self._current_filter_data(self.subfamily_filter)
            current_occ = str(self.occurrence_filter.text() or "").strip()

            self.year_filter.blockSignals(True)
            self.year_filter.clear()
            self.year_filter.addItem("Todos", "0")
            for year in years:
                self.year_filter.addItem(str(year), str(year))
            idx = self.year_filter.findData(current_year)
            self.year_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.year_filter.blockSignals(False)

            self.month_filter.blockSignals(True)
            self.month_filter.clear()
            self.month_filter.addItem("Todos", "0")
            months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            for i, label in enumerate(months, start=1):
                self.month_filter.addItem(label, str(i))
            idx = self.month_filter.findData(current_month)
            self.month_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.month_filter.blockSignals(False)

            self.manufacturer_filter.blockSignals(True)
            self.manufacturer_filter.clear()
            self.manufacturer_filter.addItem("Todos", "")
            for value, label in manufacturers:
                self.manufacturer_filter.addItem(label, value)
            idx = self.manufacturer_filter.findData(current_mfg)
            self.manufacturer_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.manufacturer_filter.blockSignals(False)

            self.family_filter.blockSignals(True)
            self.family_filter.clear()
            self.family_filter.addItem("Todas", "")
            for value, label in families:
                self.family_filter.addItem(label, value)
            idx = self.family_filter.findData(current_fam)
            self.family_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.family_filter.blockSignals(False)

            self.subfamily_filter.blockSignals(True)
            self.subfamily_filter.clear()
            self.subfamily_filter.addItem("Todas", "")
            for value, label in subfamilies:
                self.subfamily_filter.addItem(label, value)
            idx = self.subfamily_filter.findData(current_sub)
            self.subfamily_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.subfamily_filter.blockSignals(False)

            self.occurrence_filter.blockSignals(True)
            self.occurrence_filter.setText(current_occ)
            self.occurrence_filter.blockSignals(False)
        finally:
            self._building_filters = False

    def reload(self) -> None:
        if self._mode in {"in", "out", "all"} and self._building_filters:
            return

        self._movement_by_id = {}
        moves, items, manufacturer_rows, family_rows, subfamily_rows = self.movement_service.movement_payload(
            almacen_id=self._almacen_id,
            mode=self._mode,
        )

        peso_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): float(getattr(x, "articulo_envase_peso_total", 0.0) or 0.0) for x in items}
        ref_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip() for x in items}
        nombre_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip() for x in items}
        meta_by_articulo = {
            str(getattr(x, "articulo_id", "") or "").strip(): {
                "fabricante_id": str(getattr(x, "fabricante_id", "") or "").strip(),
                "familia_id": str(getattr(x, "articulo_familia_id", "") or "").strip(),
                "subfamilia_id": str(getattr(x, "articulo_subfamilia_id", "") or "").strip(),
            }
            for x in items
        }
        manufacturer_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in manufacturer_rows}
        family_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in family_rows}
        subfamily_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in subfamily_rows}

        if self._mode in {"in", "out", "all"}:
            years = sorted({mov.fecha_pedido.year for mov in moves if mov.fecha_pedido is not None}, reverse=True)
            mfg_ids = sorted({str(meta_by_articulo.get(str(getattr(m, "articulo_id", "") or "").strip(), {}).get("fabricante_id", "")) for m in moves})
            fam_ids = sorted({str(meta_by_articulo.get(str(getattr(m, "articulo_id", "") or "").strip(), {}).get("familia_id", "")) for m in moves})
            sub_ids = sorted({str(meta_by_articulo.get(str(getattr(m, "articulo_id", "") or "").strip(), {}).get("subfamilia_id", "")) for m in moves})
            self._reload_out_filters(
                years=years,
                manufacturers=[(x, manufacturer_name_by_id.get(x, x)) for x in mfg_ids if x],
                families=[(x, family_name_by_id.get(x, x)) for x in fam_ids if x],
                subfamilies=[(x, subfamily_name_by_id.get(x, x)) for x in sub_ids if x],
            )

            year_filter = int(self._current_filter_data(self.year_filter, "0") or "0")
            month_filter = int(self._current_filter_data(self.month_filter, "0") or "0")
            mfg_filter = self._current_filter_data(self.manufacturer_filter)
            fam_filter = self._current_filter_data(self.family_filter)
            sub_filter = self._current_filter_data(self.subfamily_filter)
            product_filter = str(self.occurrence_filter.text() or "").strip().lower()
            product_terms = [term for term in product_filter.split() if term]

            filtered: list[AlmacenMovimiento] = []
            for mov in moves:
                if mov.fecha_pedido is None:
                    continue
                if year_filter > 0 and mov.fecha_pedido.year != year_filter:
                    continue
                if month_filter > 0 and mov.fecha_pedido.month != month_filter:
                    continue
                art_id = str(getattr(mov, "articulo_id", "") or "").strip()
                meta = meta_by_articulo.get(art_id, {})
                if mfg_filter and str(meta.get("fabricante_id", "")) != mfg_filter:
                    continue
                if fam_filter and str(meta.get("familia_id", "")) != fam_filter:
                    continue
                if sub_filter and str(meta.get("subfamilia_id", "")) != sub_filter:
                    continue
                ref = (ref_by_articulo.get(art_id, "") or "").lower()
                nombre = (nombre_by_articulo.get(art_id, "") or "").lower()
                art_id_lower = art_id.lower()
                if product_terms:
                    searchable = " ".join([ref, nombre, art_id_lower])
                    if not all(term in searchable for term in product_terms):
                        continue
                filtered.append(mov)
            moves = filtered

        if self._mode in {"in", "out"}:
            self._movement_by_id = {
                int(getattr(mov, "id", 0) or 0): mov
                for mov in moves
                if int(getattr(mov, "id", 0) or 0) > 0
            }

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(moves))
        for row_idx, mov in enumerate(moves):
            fecha = mov.fecha_pedido.strftime("%d/%m/%Y") if mov.fecha_pedido else ""
            articulo_id = str(getattr(mov, "articulo_id", "") or "").strip()
            ref = ref_by_articulo.get(articulo_id, "") or articulo_id
            nombre = nombre_by_articulo.get(articulo_id, "") or articulo_id
            cantidad = float(getattr(mov, "cantidad", 0.0) or 0.0)
            abs_cantidad = abs(cantidad)
            kg = abs_cantidad * float(peso_by_articulo.get(articulo_id, 0.0) or 0.0)
            tipo = "Salida" if cantidad < 0 else "Entrada"

            if self._mode in {"in", "out", "all"}:
                concepto = "entrada" if cantidad > 0 else self._concept_from_albaran(str(getattr(mov, "pedido_albaran_numero", "") or "").strip())
                values = [
                    fecha,
                    ref,
                    nombre,
                    f"{'-' if self._mode == 'out' else ''}{abs_cantidad:.2f}",
                    f"{'-' if self._mode == 'out' else ''}{kg:.2f} kg",
                    str(getattr(mov, "articulo_lote", "") or "").strip(),
                    concepto,
                ]
            else:
                values = [
                    str(getattr(mov, "pedido_numero", "") or "").strip(),
                    str(getattr(mov, "pedido_albaran_numero", "") or "").strip(),
                    fecha,
                    ref,
                    nombre,
                    f"{abs_cantidad:.2f}",
                    f"{kg:.2f} kg",
                    tipo,
                ]

            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx == 0:
                    if self._mode in {"in", "out"}:
                        mov_id = int(getattr(mov, "id", 0) or 0)
                        item.setData(Qt.ItemDataRole.UserRole, {"articulo_id": articulo_id, "mov_id": mov_id})
                    else:
                        item.setData(Qt.ItemDataRole.UserRole, articulo_id)
                if (self._mode in {"in", "out"} and col_idx in (3, 4)) or (self._mode == "all" and col_idx in (3, 4, 5, 6)):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if self._mode in {"in", "out"}:
                    if col_idx in (3, 4):
                        item.setForeground(QBrush(QColor("#067647" if self._mode == "in" else "#B42318")))
                    else:
                        item.setForeground(QBrush(QColor("#000000")))
                elif self._mode == "all":
                    if col_idx in (4, 6):
                        item.setForeground(QBrush(QColor("#B42318")))
                    elif col_idx in (3, 5):
                        item.setForeground(QBrush(QColor("#067647")))
                    else:
                        item.setForeground(QBrush(QColor("#000000")))
                else:
                    if tipo == "Salida":
                        item.setForeground(QBrush(QColor("#B42318")))
                    else:
                        item.setForeground(QBrush(QColor("#067647")))
                self.table.setItem(row_idx, col_idx, item)
        if self._mode == "all":
            # Regroup by month + article + lot so entradas/salidas share line.
            grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
            for i in range(self.table.rowCount()):
                art_item = self.table.item(i, 0)
                articulo_id = str(art_item.data(Qt.ItemDataRole.UserRole) or "").strip() if art_item else ""
                fecha_txt = str(self.table.item(i, 0).text() if self.table.item(i, 0) else "")
                lote_txt = str(self.table.item(i, 5).text() if self.table.item(i, 5) else "")
                ref_txt = str(self.table.item(i, 1).text() if self.table.item(i, 1) else "")
                nombre_txt = str(self.table.item(i, 2).text() if self.table.item(i, 2) else "")
                uds_txt = str(self.table.item(i, 3).text() if self.table.item(i, 3) else "0")
                kg_txt = str(self.table.item(i, 4).text() if self.table.item(i, 4) else "0")
                concepto_txt = str(self.table.item(i, 6).text() if self.table.item(i, 6) else "")
                month_key = ""
                if fecha_txt.count("/") == 2:
                    dd, mm, yy = fecha_txt.split("/")
                    month_key = f"{yy}-{mm}"
                key = (month_key, articulo_id, lote_txt)
                row = grouped.setdefault(
                    key,
                    {
                        "mes": month_key,
                        "articulo_id": articulo_id,
                        "ref": ref_txt,
                        "nombre": nombre_txt,
                        "lote": lote_txt,
                        "entradas_uds": 0.0,
                        "salidas_uds": 0.0,
                        "entradas_kg": 0.0,
                        "salidas_kg": 0.0,
                    },
                )
                uds = float(uds_txt.replace(",", ".").replace("-", "") or 0.0)
                kg_clean = kg_txt.lower().replace("kg", "").replace(",", ".").strip()
                kg = float(kg_clean.replace("-", "") or 0.0)
                if concepto_txt == "entrada":
                    row["entradas_uds"] += uds
                    row["entradas_kg"] += kg
                else:
                    row["salidas_uds"] += uds
                    row["salidas_kg"] += kg

            self.table.setSortingEnabled(False)
            self.table.setRowCount(len(grouped))
            for row_idx, values in enumerate(grouped.values()):
                cols = [
                    values["mes"],
                    values["ref"],
                    values["nombre"],
                    f'{values["entradas_uds"]:.2f}',
                    f'-{values["salidas_uds"]:.2f}' if values["salidas_uds"] else "0.00",
                    f'{values["entradas_kg"]:.2f} kg',
                    f'-{values["salidas_kg"]:.2f} kg' if values["salidas_kg"] else "0.00 kg",
                    values["lote"],
                ]
                for col_idx, val in enumerate(cols):
                    it = QTableWidgetItem(str(val))
                    if col_idx == 0:
                        it.setData(Qt.ItemDataRole.UserRole, values["articulo_id"])
                    if col_idx in (3, 4, 5, 6):
                        it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    if col_idx in (4, 6):
                        it.setForeground(QBrush(QColor("#B42318")))
                    elif col_idx in (3, 5):
                        it.setForeground(QBrush(QColor("#067647")))
                    else:
                        it.setForeground(QBrush(QColor("#000000")))
                    self.table.setItem(row_idx, col_idx, it)
        self.table.setSortingEnabled(True)

    def selected_articulo_id_from_row(self, row: int) -> str:
        item = self.table.item(row, 0)
        if item is None:
            return ""
        data = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(data, dict):
            return str(data.get("articulo_id") or "").strip()
        return str(data or "").strip()

    def _selected_movement(self) -> AlmacenMovimiento | None:
        selected = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected:
            return None
        row = selected[0].row()
        item = self.table.item(row, 0)
        if item is None:
            return None
        data = item.data(Qt.ItemDataRole.UserRole)
        mov_id = 0
        if isinstance(data, dict):
            mov_id = int(data.get("mov_id") or 0)
        if mov_id <= 0:
            return None
        return self._movement_by_id.get(mov_id)

    def _manual_schema(self) -> list[dict[str, Any]]:
        return [
            {"name": "fecha_pedido", "label": "Fecha (YYYY-MM-DD)"},
            {"name": "articulo_id", "label": "Articulo_ID"},
            {"name": "cantidad", "label": "Cantidad (uds)"},
            {"name": "articulo_lote", "label": "Lote"},
            {"name": "articulo_caducidad", "label": "Caducidad (YYYY-MM-DD)"},
            {"name": "motivo", "label": "Motivo"},
            {"name": "usuario", "label": "Usuario"},
            {"name": "observacion", "label": "Observacion"},
        ]

    def _parse_date_field(self, text: str, *, allow_empty: bool = False) -> date | None:
        raw = str(text or "").strip()
        if not raw:
            return None if allow_empty else date.today()
        try:
            if "/" in raw and raw.count("/") == 2:
                dd, mm, yy = raw.split("/")
                return date(int(yy), int(mm), int(dd))
            return date.fromisoformat(raw)
        except Exception:
            raise ValueError(f"Fecha no valida: {raw}") from None

    def _build_manual_albaran(self, *, motivo: str, usuario: str, observacion: str) -> str:
        mot = str(motivo or "").strip()
        usr = str(usuario or "").strip()
        obs = str(observacion or "").strip().replace("|", "/")
        base = f"MANUAL-{self._mode.upper()}|MOT:{mot}|USR:{usr}"
        if obs:
            base = f"{base}|OBS:{obs}"
        return base[:255]

    def _current_stock_for(self, *, articulo_id: str, lote: str, caducidad: date | None) -> float:
        return self.movement_service.current_stock_for(
            almacen_id=self._almacen_id,
            articulo_id=articulo_id,
            lote=lote,
            caducidad=caducidad,
        )

    def _create_manual_move(self) -> None:
        article_options = self._load_article_options()
        if not article_options:
            QMessageBox.warning(self, "Movimiento manual", "No hay artículos disponibles para seleccionar.")
            return
        initial = {"fecha_pedido": date.today().isoformat(), "cantidad": 1.0}
        dialog = ManualMovementDialog(
            title=f"Nueva manual ({'Entrada' if self._mode == 'in' else 'Salida'})",
            article_options=article_options,
            initial=initial,
            parent=self,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        payload = dialog.payload()
        try:
            self._save_manual_move(payload, existing=None)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Movimiento manual", str(exc))
            return
        self.reload()

    def _edit_manual_move(self) -> None:
        mov = self._selected_movement()
        if mov is None:
            QMessageBox.warning(self, "Movimiento manual", "Selecciona un movimiento.")
            return
        albaran = str(getattr(mov, "pedido_albaran_numero", "") or "").strip()
        if not albaran.startswith("MANUAL-"):
            QMessageBox.warning(self, "Movimiento manual", "Solo se pueden editar movimientos manuales.")
            return
        motivo, usuario, observacion = self._parse_manual_albaran(albaran)
        initial = {
            "fecha_pedido": getattr(mov, "fecha_pedido", date.today()).isoformat(),
            "articulo_id": str(getattr(mov, "articulo_id", "") or "").strip(),
            "cantidad": abs(float(getattr(mov, "cantidad", 0.0) or 0.0)),
            "articulo_lote": str(getattr(mov, "articulo_lote", "") or "").strip(),
            "articulo_caducidad": getattr(mov, "articulo_caducidad").isoformat() if getattr(mov, "articulo_caducidad", None) else "",
            "motivo": motivo,
            "usuario": usuario,
            "observacion": observacion,
        }
        article_options = self._load_article_options()
        if not article_options:
            QMessageBox.warning(self, "Movimiento manual", "No hay artículos disponibles para seleccionar.")
            return
        dialog = ManualMovementDialog(
            title=f"Editar manual ({'Entrada' if self._mode == 'in' else 'Salida'})",
            article_options=article_options,
            initial=initial,
            parent=self,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        payload = dialog.payload()
        try:
            self._save_manual_move(payload, existing=mov)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Movimiento manual", str(exc))
            return
        self.reload()

    def _load_article_options(self) -> list[tuple[str, str, str]]:
        return self.movement_service.article_options(self._almacen_id)

    def _parse_manual_albaran(self, albaran: str) -> tuple[str, str, str]:
        motivo = ""
        usuario = ""
        observacion = ""
        for part in str(albaran or "").split("|"):
            part = part.strip()
            if part.startswith("MOT:"):
                motivo = part[4:].strip()
            elif part.startswith("USR:"):
                usuario = part[4:].strip()
            elif part.startswith("OBS:"):
                observacion = part[4:].strip()
        return motivo, usuario, observacion

    def _reverse_manual_move(self) -> None:
        mov = self._selected_movement()
        if mov is None:
            QMessageBox.warning(self, "Movimiento manual", "Selecciona un movimiento.")
            return
        albaran = str(getattr(mov, "pedido_albaran_numero", "") or "").strip()
        if not albaran.startswith("MANUAL-"):
            QMessageBox.warning(self, "Movimiento manual", "Solo se pueden anular movimientos manuales.")
            return
        answer = QMessageBox.question(self, "Confirmar", "Se creará un contramovimiento para anular este ajuste manual. ¿Continuar?")
        if answer != QMessageBox.StandardButton.Yes:
            return
        self.movement_service.reverse_manual_move(mov)
        self.reload()

    def _save_manual_move_legacy(self, payload: dict[str, Any], existing: AlmacenMovimiento | None) -> None:
        articulo_id = str(payload.get("articulo_id") or "").strip()
        if not articulo_id:
            raise ValueError("Articulo_ID es obligatorio.")
        cantidad_raw = str(payload.get("cantidad") or "").strip().replace(",", ".")
        try:
            cantidad_abs = abs(float(cantidad_raw))
        except Exception:
            raise ValueError("Cantidad no valida.") from None
        if cantidad_abs <= 0:
            raise ValueError("La cantidad debe ser mayor que 0.")
        fecha_pedido = self._parse_date_field(str(payload.get("fecha_pedido") or ""), allow_empty=False) or date.today()
        cad = self._parse_date_field(str(payload.get("articulo_caducidad") or ""), allow_empty=True)
        lote = str(payload.get("articulo_lote") or "").strip()
        motivo = str(payload.get("motivo") or "").strip()
        usuario = str(payload.get("usuario") or "").strip()
        observacion = str(payload.get("observacion") or "").strip()
        if not motivo:
            raise ValueError("Motivo es obligatorio.")
        if not usuario:
            raise ValueError("Usuario es obligatorio.")
        cantidad_signed = cantidad_abs if self._mode == "in" else -cantidad_abs
        if self._mode == "out":
            stock = self._current_stock_for(articulo_id=articulo_id, lote=lote, caducidad=cad)
            if existing is not None:
                stock -= float(getattr(existing, "cantidad", 0.0) or 0.0)
            if stock + cantidad_signed < -0.0001:
                raise ValueError("La salida manual dejaría stock negativo para ese producto/lote.")
        albaran = self._build_manual_albaran(motivo=motivo, usuario=usuario, observacion=observacion)
        return self.movement_service.save_manual_move(
            payload=payload,
            mode=self._mode,
            almacen_id=self._almacen_id,
            existing=existing,
            fecha_pedido=fecha_pedido,
            caducidad=cad,
            albaran=albaran,
        )


    def _save_manual_move(self, payload: dict[str, Any], existing: AlmacenMovimiento | None) -> AlmacenMovimiento:
        context = self.manual_move_flow_service.build_manual_move_context(
            payload,
            mode=self._mode,
            almacen_id=self._almacen_id,
            existing=existing,
        )
        if context.status != "ready":
            raise ValueError(context.message or "No se pudo guardar el movimiento manual.")
        result = self.manual_move_flow_service.submit_manual_move(context)
        if result.status != "success" or result.move is None:
            raise ValueError(result.message or "No se pudo guardar el movimiento manual.")
        return result.move


def _compute_current_stock_rows(moves: list[AlmacenMovimiento]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, date | None], dict[str, Any]] = {}
    for mov in moves:
        articulo_id = str(getattr(mov, "articulo_id", "") or "").strip()
        if not articulo_id:
            continue
        lote = str(getattr(mov, "articulo_lote", "") or "").strip()
        cad = getattr(mov, "articulo_caducidad", None)
        key = (articulo_id, lote, cad)
        row = grouped.setdefault(
            key,
            {
                "articulo_id": articulo_id,
                "lote": lote,
                "caducidad": cad,
                "cantidad": 0.0,
                "last_date": None,
            },
        )
        row["cantidad"] = float(row["cantidad"]) + float(getattr(mov, "cantidad", 0.0) or 0.0)
        mov_date = getattr(mov, "fecha_pedido", None)
        if mov_date is not None and (row["last_date"] is None or mov_date > row["last_date"]):
            row["last_date"] = mov_date
    return [row for row in grouped.values() if float(row.get("cantidad", 0.0) or 0.0) > 0]


class ManualMovementDialog(QDialog):
    def __init__(
        self,
        *,
        title: str,
        article_options: list[tuple[str, str, str]],
        initial: dict[str, Any] | None = None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(520, 320)
        self._article_options = article_options
        self._initial = initial or {}
        self._build_ui()
        self._load_initial()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        form.setFormAlignment(Qt.AlignmentFlag.AlignTop)
        form.setVerticalSpacing(10)

        self.selected_articulo_id = ""
        art_row = QHBoxLayout()
        self.ref_input = QLineEdit()
        self.ref_input.setReadOnly(True)
        self.ref_input.setText("Ref")
        self.select_article_btn = QPushButton("Seleccionar...")
        base_h = self.article_name.sizeHint().height() if hasattr(self, "article_name") else self.ref_input.sizeHint().height()
        base_h = max(int(base_h), 32)
        self.ref_input.setMinimumHeight(base_h)
        self.select_article_btn.setMinimumHeight(base_h)
        self.select_article_btn.setStyleSheet(
            "QPushButton { background-color: #1E4FA1; color: white; border: 1px solid #1E4FA1; border-radius: 4px; padding: 0 12px; }"
            "QPushButton:hover { background-color: #245dbd; border-color: #245dbd; }"
            "QPushButton:pressed { background-color: #18448b; border-color: #18448b; }"
        )
        self.select_article_btn.clicked.connect(self._pick_article)
        art_row.setContentsMargins(0, 0, 0, 0)
        art_row.setSpacing(8)
        art_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        art_row.addWidget(self.ref_input, 1)
        art_row.addWidget(self.select_article_btn)
        art_wrap = QWidget()
        art_wrap.setLayout(art_row)
        art_wrap.setMinimumHeight(base_h)
        art_label = QLabel("Artículo (ref)")
        art_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        art_label.setMinimumHeight(base_h)
        form.addRow(art_label, art_wrap)

        self.article_name = QLineEdit()
        self.article_name.setReadOnly(True)
        form.addRow("Nombre artículo", self.article_name)
        # Ajusta la altura del ref al estándar del resto de QLineEdit del formulario.
        self.ref_input.setFixedHeight(self.article_name.sizeHint().height())

        self.qty_input = QDoubleSpinBox()
        self.qty_input.setRange(0.01, 99999999.0)
        self.qty_input.setDecimals(2)
        self.qty_input.setAlignment(Qt.AlignmentFlag.AlignRight)
        form.addRow("Cantidad (uds)", self.qty_input)

        self.lote_input = QLineEdit()
        form.addRow("Lote", self.lote_input)

        self.fecha_input = QDateEdit()
        self.fecha_input.setCalendarPopup(True)
        self.fecha_input.setDisplayFormat("dd-MM-yyyy")
        self.fecha_input.setDate(QDate.currentDate())
        form.addRow("Fecha", self.fecha_input)

        cad_row = QHBoxLayout()
        self.cad_input = QDateEdit()
        self.cad_input.setCalendarPopup(True)
        self.cad_input.setDisplayFormat("dd-MM-yyyy")
        self.cad_input.setDate(QDate.currentDate())
        self.no_cad_check = QCheckBox("Sin caducidad")
        cad_h = max(self.cad_input.sizeHint().height(), self.no_cad_check.sizeHint().height())
        self.cad_input.setFixedHeight(cad_h)
        self.no_cad_check.setFixedHeight(cad_h)
        self.no_cad_check.setChecked(True)
        self.no_cad_check.toggled.connect(lambda checked: self.cad_input.setEnabled(not checked))
        self.cad_input.setEnabled(False)
        cad_row.setContentsMargins(0, 0, 0, 0)
        cad_row.setSpacing(8)
        cad_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        cad_row.addWidget(self.cad_input, 1)
        cad_row.addWidget(self.no_cad_check)
        cad_wrap = QWidget()
        cad_wrap.setLayout(cad_row)
        form.addRow("Caducidad", cad_wrap)

        self.motivo_input = QLineEdit()
        form.addRow("Motivo", self.motivo_input)
        self.usuario_input = QLineEdit()
        form.addRow("Usuario", self.usuario_input)
        self.obs_input = QLineEdit()
        form.addRow("Observación", self.obs_input)

        layout.addLayout(form)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        save_btn = buttons.button(QDialogButtonBox.StandardButton.Save)
        cancel_btn = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        if save_btn is not None:
            save_btn.setText("Guardar")
        if cancel_btn is not None:
            cancel_btn.setText("Cancelar")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _set_article(self, articulo_id: str) -> None:
        selected = next((opt for opt in self._article_options if opt[0] == articulo_id), None)
        if selected is None:
            self.selected_articulo_id = ""
            self.ref_input.setText("Ref")
            self.article_name.setText("")
            return
        self.selected_articulo_id = str(selected[0] or "").strip()
        self.ref_input.setText(str(selected[1] or "").strip())
        self.article_name.setText(str(selected[2] or "").strip())

    def _pick_article(self) -> None:
        picker = ArticlePickerDialog(self._article_options, parent=self)
        if picker.exec() != QDialog.DialogCode.Accepted:
            return
        articulo_id = picker.selected_articulo_id()
        if articulo_id:
            self._set_article(articulo_id)

    def _load_initial(self) -> None:
        article_id = str(self._initial.get("articulo_id") or "").strip()
        if article_id:
            self._set_article(article_id)
        try:
            self.qty_input.setValue(float(self._initial.get("cantidad") or 1.0))
        except Exception:
            self.qty_input.setValue(1.0)
        self.lote_input.setText(str(self._initial.get("articulo_lote") or "").strip())
        fecha_txt = str(self._initial.get("fecha_pedido") or "").strip()
        if fecha_txt:
            parsed = QDate.fromString(fecha_txt, "yyyy-MM-dd")
            if not parsed.isValid():
                parsed = QDate.fromString(fecha_txt, "dd-MM-yyyy")
            if parsed.isValid():
                self.fecha_input.setDate(parsed)
        cad_txt = str(self._initial.get("articulo_caducidad") or "").strip()
        if cad_txt:
            parsed = QDate.fromString(cad_txt, "yyyy-MM-dd")
            if not parsed.isValid():
                parsed = QDate.fromString(cad_txt, "dd-MM-yyyy")
            if parsed.isValid():
                self.no_cad_check.setChecked(False)
                self.cad_input.setEnabled(True)
                self.cad_input.setDate(parsed)
        self.motivo_input.setText(str(self._initial.get("motivo") or "").strip())
        self.usuario_input.setText(str(self._initial.get("usuario") or "").strip())
        self.obs_input.setText(str(self._initial.get("observacion") or "").strip())

    def payload(self) -> dict[str, Any]:
        cad = "" if self.no_cad_check.isChecked() else self.cad_input.date().toString("yyyy-MM-dd")
        return {
            "articulo_id": str(self.selected_articulo_id or "").strip(),
            "cantidad": float(self.qty_input.value()),
            "articulo_lote": str(self.lote_input.text() or "").strip(),
            "fecha_pedido": self.fecha_input.date().toString("yyyy-MM-dd"),
            "articulo_caducidad": cad,
            "motivo": str(self.motivo_input.text() or "").strip(),
            "usuario": str(self.usuario_input.text() or "").strip(),
            "observacion": str(self.obs_input.text() or "").strip(),
        }


class ArticlePickerDialog(QDialog):
    def __init__(self, article_options: list[tuple[str, str, str]], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Seleccionar artículo")
        self.setModal(True)
        self.resize(720, 420)
        self._options = article_options
        self._build_ui()
        self._reload_rows()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Filtrar por ref, nombre o id...")
        self.filter_input.textChanged.connect(self._reload_rows)
        layout.addWidget(self.filter_input)

        self.table = QTableWidget(0, 2)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setHorizontalHeaderLabels(["Ref", "Nombre"])
        self.table.setColumnWidth(0, 180)
        self.table.cellDoubleClicked.connect(lambda _r, _c: self.accept())
        layout.addWidget(self.table, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        ok_btn = buttons.button(QDialogButtonBox.StandardButton.Ok)
        cancel_btn = buttons.button(QDialogButtonBox.StandardButton.Cancel)
        if ok_btn is not None:
            ok_btn.setText("Seleccionar")
        if cancel_btn is not None:
            cancel_btn.setText("Cancelar")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _reload_rows(self) -> None:
        terms = [x for x in str(self.filter_input.text() or "").strip().lower().split() if x]
        rows: list[tuple[str, str, str]] = []
        for articulo_id, ref, nombre in self._options:
            haystack = " ".join([str(articulo_id or "").lower(), str(ref or "").lower(), str(nombre or "").lower()])
            if terms and not all(term in haystack for term in terms):
                continue
            rows.append((str(articulo_id or "").strip(), str(ref or "").strip(), str(nombre or "").strip()))
        self.table.setRowCount(len(rows))
        for i, (art_id, ref, nombre) in enumerate(rows):
            ref_item = QTableWidgetItem(ref)
            ref_item.setData(Qt.ItemDataRole.UserRole, art_id)
            self.table.setItem(i, 0, ref_item)
            self.table.setItem(i, 1, QTableWidgetItem(nombre))
        if rows:
            self.table.selectRow(0)

    def selected_articulo_id(self) -> str:
        selected = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not selected:
            return ""
        row = selected[0].row()
        item = self.table.item(row, 0)
        return str(item.data(Qt.ItemDataRole.UserRole) or "").strip() if item else ""


class StockTab(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._almacen_id = ""
        self._building_filters = False
        self.movement_service = WarehouseMovementService()
        self.settings = WarehouseSettingsService()
        loaded = self.settings.load()
        self.low_stock_threshold_units = float(loaded.get("low_stock_threshold_units") or 1.0)
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        filters = QHBoxLayout()

        filters.addWidget(QLabel("Fabricante"))
        self.manufacturer_filter = QComboBox()
        self.manufacturer_filter.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.manufacturer_filter)

        filters.addWidget(QLabel("Familia"))
        self.family_filter = QComboBox()
        self.family_filter.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.family_filter)

        filters.addWidget(QLabel("Subfamilia"))
        self.subfamily_filter = QComboBox()
        self.subfamily_filter.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.subfamily_filter)

        filters.addWidget(QLabel("Riesgo"))
        self.risk_filter = QComboBox()
        self.risk_filter.addItem("Todos", "all")
        self.risk_filter.addItem("Caducado", "expired")
        self.risk_filter.addItem("Caduca <= 30 días", "soon")
        self.risk_filter.addItem("Bajo stock", "low")
        self.risk_filter.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.risk_filter)

        filters.addWidget(QLabel("Umbral bajo stock (uds)"))
        self.low_stock_spin = QDoubleSpinBox()
        self.low_stock_spin.setDecimals(2)
        self.low_stock_spin.setRange(0.0, 999999.0)
        self.low_stock_spin.setSingleStep(0.5)
        self.low_stock_spin.setValue(self.low_stock_threshold_units)
        self.low_stock_spin.valueChanged.connect(self._on_low_stock_changed)
        filters.addWidget(self.low_stock_spin)

        filters.addStretch(1)
        layout.addLayout(filters)
        product_filters = QHBoxLayout()
        product_filters.addWidget(QLabel("Producto"))
        self.product_filter = QLineEdit()
        self.product_filter.setPlaceholderText("Nombre o ref...")
        self.product_filter.setMinimumWidth(420)
        self.product_filter.textChanged.connect(self.reload)
        product_filters.addWidget(self.product_filter)
        product_filters.addStretch(1)
        layout.addLayout(product_filters)

        self.table = QTableWidget(0, 9)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setStyleSheet("QTableWidget::item:focus { border: none; outline: 0; }")
        header = self.table.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Fixed)
        self.table.setHorizontalHeaderLabels(["Ref.", "Nombre", "Lote", "Caduca", "Días", "Stock Uds", "Stock Kg", "Últ. mov.", "Estado"])
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 110)
        self.table.setColumnWidth(4, 70)
        self.table.setColumnWidth(5, 90)
        self.table.setColumnWidth(6, 95)
        self.table.setColumnWidth(7, 110)
        self.table.setColumnWidth(8, 120)
        layout.addWidget(self.table, 1)

    def set_almacen_filter(self, almacen_id: str) -> None:
        self._almacen_id = str(almacen_id or "").strip()
        self.reload()

    def _current_filter_data(self, combo: QComboBox, default: str = "") -> str:
        return str(combo.currentData() or default).strip()

    def _reload_filters(
        self,
        manufacturers: list[tuple[str, str]],
        families: list[tuple[str, str]],
        subfamilies: list[tuple[str, str]],
    ) -> None:
        self._building_filters = True
        try:
            current_mfg = self._current_filter_data(self.manufacturer_filter)
            current_fam = self._current_filter_data(self.family_filter)
            current_sub = self._current_filter_data(self.subfamily_filter)
            current_product = str(self.product_filter.text() or "").strip()

            self.manufacturer_filter.blockSignals(True)
            self.manufacturer_filter.clear()
            self.manufacturer_filter.addItem("Todos", "")
            for value, label in manufacturers:
                self.manufacturer_filter.addItem(label, value)
            idx = self.manufacturer_filter.findData(current_mfg)
            self.manufacturer_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.manufacturer_filter.blockSignals(False)

            self.family_filter.blockSignals(True)
            self.family_filter.clear()
            self.family_filter.addItem("Todas", "")
            for value, label in families:
                self.family_filter.addItem(label, value)
            idx = self.family_filter.findData(current_fam)
            self.family_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.family_filter.blockSignals(False)

            self.subfamily_filter.blockSignals(True)
            self.subfamily_filter.clear()
            self.subfamily_filter.addItem("Todas", "")
            for value, label in subfamilies:
                self.subfamily_filter.addItem(label, value)
            idx = self.subfamily_filter.findData(current_sub)
            self.subfamily_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.subfamily_filter.blockSignals(False)

            self.product_filter.blockSignals(True)
            self.product_filter.setText(current_product)
            self.product_filter.blockSignals(False)
        finally:
            self._building_filters = False

    def reload(self) -> None:
        if self._building_filters:
            return
        moves, items, manufacturer_rows, family_rows, subfamily_rows = self.movement_service.movement_payload(
            almacen_id=self._almacen_id,
            mode="all",
        )

        stock_rows = _compute_current_stock_rows(moves)
        peso_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): float(getattr(x, "articulo_envase_peso_total", 0.0) or 0.0) for x in items}
        ref_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip() for x in items}
        nombre_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip() for x in items}
        meta_by_articulo = {
            str(getattr(x, "articulo_id", "") or "").strip(): {
                "fabricante_id": str(getattr(x, "fabricante_id", "") or "").strip(),
                "familia_id": str(getattr(x, "articulo_familia_id", "") or "").strip(),
                "subfamilia_id": str(getattr(x, "articulo_subfamilia_id", "") or "").strip(),
            }
            for x in items
        }
        manufacturer_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in manufacturer_rows}
        family_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in family_rows}
        subfamily_name_by_id = {str(i or "").strip(): (str(n or "").strip() or str(i or "").strip()) for i, n in subfamily_rows}

        mfg_ids = sorted({str(meta_by_articulo.get(str(r.get("articulo_id", "")), {}).get("fabricante_id", "")) for r in stock_rows})
        fam_ids = sorted({str(meta_by_articulo.get(str(r.get("articulo_id", "")), {}).get("familia_id", "")) for r in stock_rows})
        sub_ids = sorted({str(meta_by_articulo.get(str(r.get("articulo_id", "")), {}).get("subfamilia_id", "")) for r in stock_rows})
        self._reload_filters(
            manufacturers=[(x, manufacturer_name_by_id.get(x, x)) for x in mfg_ids if x],
            families=[(x, family_name_by_id.get(x, x)) for x in fam_ids if x],
            subfamilies=[(x, subfamily_name_by_id.get(x, x)) for x in sub_ids if x],
        )

        mfg_filter = self._current_filter_data(self.manufacturer_filter)
        fam_filter = self._current_filter_data(self.family_filter)
        sub_filter = self._current_filter_data(self.subfamily_filter)
        risk_filter = self._current_filter_data(self.risk_filter, "all")
        product_terms = [term for term in str(self.product_filter.text() or "").strip().lower().split() if term]
        low_threshold = float(self.low_stock_spin.value() if hasattr(self, "low_stock_spin") else self.low_stock_threshold_units)

        today = date.today()
        soon_limit = today + timedelta(days=30)
        filtered_rows: list[dict[str, Any]] = []
        for row in stock_rows:
            art_id = str(row.get("articulo_id", "") or "").strip()
            meta = meta_by_articulo.get(art_id, {})
            if mfg_filter and str(meta.get("fabricante_id", "")) != mfg_filter:
                continue
            if fam_filter and str(meta.get("familia_id", "")) != fam_filter:
                continue
            if sub_filter and str(meta.get("subfamilia_id", "")) != sub_filter:
                continue

            ref = (ref_by_articulo.get(art_id, "") or "").lower()
            nombre = (nombre_by_articulo.get(art_id, "") or "").lower()
            if product_terms:
                searchable = " ".join([ref, nombre, art_id.lower()])
                if not all(term in searchable for term in product_terms):
                    continue

            cad = row.get("caducidad")
            qty = float(row.get("cantidad", 0.0) or 0.0)
            if cad is not None and cad < today:
                state = "Caducado"
            elif cad is not None and cad <= soon_limit:
                state = "Caduca pronto"
            elif qty < low_threshold:
                state = "Bajo stock"
            else:
                state = "OK"

            if risk_filter == "expired" and state != "Caducado":
                continue
            if risk_filter == "soon" and state != "Caduca pronto":
                continue
            if risk_filter == "low" and state != "Bajo stock":
                continue

            row["estado"] = state
            filtered_rows.append(row)

        state_order = {"Caducado": 0, "Caduca pronto": 1, "Bajo stock": 2, "OK": 3}
        filtered_rows.sort(
            key=lambda r: (
                state_order.get(str(r.get("estado", "OK")), 9),
                r.get("caducidad") or date.max,
                str(nombre_by_articulo.get(str(r.get("articulo_id", "")), "")),
            )
        )

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(filtered_rows))
        for row_idx, row in enumerate(filtered_rows):
            art_id = str(row.get("articulo_id", "") or "").strip()
            ref = ref_by_articulo.get(art_id, "") or art_id
            nombre = nombre_by_articulo.get(art_id, "") or art_id
            cad = row.get("caducidad")
            qty = float(row.get("cantidad", 0.0) or 0.0)
            kg = qty * float(peso_by_articulo.get(art_id, 0.0) or 0.0)
            last_date = row.get("last_date")
            dias = ""
            if cad is not None:
                dias = str((cad - today).days)
            values = [
                ref,
                nombre,
                str(row.get("lote", "") or ""),
                cad.strftime("%d/%m/%Y") if cad else "",
                dias,
                f"{qty:.2f}",
                f"{kg:.2f} kg",
                last_date.strftime("%d/%m/%Y") if last_date else "",
                str(row.get("estado", "OK")),
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx == 0:
                    item.setData(Qt.ItemDataRole.UserRole, art_id)
                if col_idx in (4, 5, 6):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if col_idx == 8:
                    state = str(row.get("estado", "OK"))
                    if state == "Caducado":
                        item.setForeground(QBrush(QColor("#B42318")))
                    elif state == "Caduca pronto":
                        item.setForeground(QBrush(QColor("#EF6C00")))
                    elif state == "Bajo stock":
                        item.setForeground(QBrush(QColor("#7A5AF8")))
                    else:
                        item.setForeground(QBrush(QColor("#067647")))
                self.table.setItem(row_idx, col_idx, item)
        self.table.setSortingEnabled(True)

    def _on_low_stock_changed(self, value: float) -> None:
        self.low_stock_threshold_units = max(0.0, float(value or 0.0))
        try:
            self.settings.save(self.low_stock_threshold_units)
        except Exception:
            pass
        self.reload()


class InventariosTab(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._almacen_id = ""
        self._pending_ajustes: list[dict[str, Any]] = []
        self.inventory_service = WarehouseInventoryService()
        self.inventory_adjustment_preparation_service = WarehouseInventoryAdjustmentPreparationService()
        self.view = InventoryTabView()
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        self.inner_tabs = QTabWidget()
        conteo_tab = QWidget()
        historial_tab = QWidget()
        conteo_layout = QVBoxLayout(conteo_tab)
        historial_layout = QVBoxLayout(historial_tab)

        top = QHBoxLayout()
        top.addWidget(QLabel(self.view.intro_text))
        top.addStretch(1)
        export_btn = QPushButton(self.view.export_template_button_label)
        export_btn.setProperty("btnRole", "secondary")
        export_btn.clicked.connect(self._export_count_template)
        import_btn = QPushButton(self.view.import_count_button_label)
        import_btn.setProperty("btnRole", "secondary")
        import_btn.clicked.connect(self._import_count_template)
        refresh_btn = QPushButton(self.view.refresh_button_label)
        refresh_btn.setProperty("btnRole", "secondary")
        refresh_btn.clicked.connect(self.reload)
        prepare_btn = QPushButton(self.view.prepare_adjustments_button_label)
        prepare_btn.setProperty("btnRole", "warning")
        prepare_btn.clicked.connect(self._prepare_adjustments)
        apply_btn = QPushButton(self.view.apply_adjustments_button_label)
        apply_btn.setProperty("btnRole", "warning")
        apply_btn.clicked.connect(self._apply_adjustments)
        top.addWidget(export_btn)
        top.addWidget(import_btn)
        top.addWidget(refresh_btn)
        top.addWidget(prepare_btn)
        top.addWidget(apply_btn)
        conteo_layout.addLayout(top)

        approval = QHBoxLayout()
        approval.addWidget(QLabel(self.view.search_label))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(self.view.search_placeholder)
        self.search_input.textChanged.connect(self.reload)
        approval.addWidget(self.search_input)
        approval.addWidget(QLabel(self.view.counter_label))
        self.counter_input = QLineEdit()
        self.counter_input.setPlaceholderText(self.view.counter_placeholder)
        approval.addWidget(self.counter_input)
        approval.addWidget(QLabel(self.view.approver_label))
        self.approver_input = QLineEdit()
        self.approver_input.setPlaceholderText(self.view.approver_placeholder)
        approval.addWidget(self.approver_input)
        self.pending_label = QLabel(self.view.pending_label)
        approval.addWidget(self.pending_label)
        approval.addStretch(1)
        conteo_layout.addLayout(approval)

        self.table = QTableWidget(0, 8)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setStyleSheet("QTableWidget::item:focus { border: none; outline: 0; }")
        header = self.table.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.table.setHorizontalHeaderLabels(["Ref.", "Nombre", "Lote", "Caduca", "Te?rico Uds", "Conteo Uds", "Diferencia", "Kg ajuste"])
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 110)
        self.table.setColumnWidth(4, 95)
        self.table.setColumnWidth(5, 95)
        self.table.setColumnWidth(6, 90)
        self.table.setColumnWidth(7, 95)
        self.table.itemChanged.connect(self._on_item_changed)
        conteo_layout.addWidget(self.table, 1)

        history_top = QHBoxLayout()
        history_label = QLabel(self.view.history_title)
        history_label.setProperty("role", "sectionTitle")
        history_top.addWidget(history_label)
        history_top.addStretch(1)
        export_history_btn = QPushButton(self.view.export_history_button_label)
        export_history_btn.setProperty("btnRole", "secondary")
        export_history_btn.clicked.connect(self._export_history_excel)
        history_top.addWidget(export_history_btn)
        historial_layout.addLayout(history_top)

        self.history_table = QTableWidget(0, 7)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.history_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.history_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.history_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.history_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.history_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.history_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.history_table.setHorizontalHeaderLabels(["C?digo", "Fecha", "Contador", "Aprobador", "L?neas", "Ajustes", "Estado"])
        self.history_table.setColumnWidth(0, 130)
        self.history_table.setColumnWidth(1, 110)
        self.history_table.setColumnWidth(4, 70)
        self.history_table.setColumnWidth(5, 70)
        self.history_table.setColumnWidth(6, 90)
        self.history_table.itemSelectionChanged.connect(self._reload_history_detail)
        historial_layout.addWidget(self.history_table, 1)

        self.history_detail_table = QTableWidget(0, 8)
        self.history_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_detail_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.history_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_detail_table.verticalHeader().setVisible(False)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.history_detail_table.setHorizontalHeaderLabels(["Ref.", "Nombre", "Lote", "Caduca", "Te?rico", "Conteo", "Diferencia", "Kg ajuste"])
        self.history_detail_table.setColumnWidth(0, 110)
        self.history_detail_table.setColumnWidth(2, 110)
        self.history_detail_table.setColumnWidth(3, 110)
        self.history_detail_table.setColumnWidth(4, 90)
        self.history_detail_table.setColumnWidth(5, 90)
        self.history_detail_table.setColumnWidth(6, 90)
        self.history_detail_table.setColumnWidth(7, 95)
        historial_layout.addWidget(self.history_detail_table, 1)

        self.inner_tabs.addTab(conteo_tab, "Conteo")
        self.inner_tabs.addTab(historial_tab, "Historial")
        layout.addWidget(self.inner_tabs, 1)

    def set_almacen_filter(self, almacen_id: str) -> None:
        self._almacen_id = str(almacen_id or "").strip()
        self.reload()

    def reload(self) -> None:
        payload = self.inventory_service.stock_payload(self._almacen_id)
        moves = payload.moves
        items = payload.items
        rows = _compute_current_stock_rows(moves)
        ref_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip() for x in items}
        nombre_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip() for x in items}
        peso_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): float(getattr(x, "articulo_envase_peso_total", 0.0) or 0.0) for x in items}

        rows.sort(key=lambda r: (str(ref_by_articulo.get(str(r.get("articulo_id", "")), "")), str(r.get("lote", ""))))
        search_terms = [x for x in str(self.search_input.text() or "").strip().lower().split() if x]
        if search_terms:
            filtered_rows: list[dict[str, Any]] = []
            for row in rows:
                art_id = str(row.get("articulo_id", "") or "").strip()
                searchable = " ".join(
                    [
                        str(ref_by_articulo.get(art_id, "") or "").lower(),
                        str(nombre_by_articulo.get(art_id, "") or "").lower(),
                        str(row.get("lote", "") or "").lower(),
                        art_id.lower(),
                    ]
                )
                if all(term in searchable for term in search_terms):
                    filtered_rows.append(row)
            rows = filtered_rows
        self.table.blockSignals(True)
        self.table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            art_id = str(row.get("articulo_id", "") or "").strip()
            qty = float(row.get("cantidad", 0.0) or 0.0)
            cad = row.get("caducidad")
            values = [
                ref_by_articulo.get(art_id, "") or art_id,
                nombre_by_articulo.get(art_id, "") or art_id,
                str(row.get("lote", "") or ""),
                cad.strftime("%d/%m/%Y") if cad else "",
                f"{qty:.2f}",
                "",
                "0.00",
                "0.00 kg",
            ]
            for col_idx, value in enumerate(values):
                it = QTableWidgetItem(value)
                if col_idx == 0:
                    it.setData(Qt.ItemDataRole.UserRole, (art_id, str(row.get("lote", "") or ""), cad.isoformat() if cad else ""))
                if col_idx in (4, 5, 6, 7):
                    it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if col_idx != 5:
                    it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row_idx, col_idx, it)
            self.table.item(row_idx, 5).setData(Qt.ItemDataRole.UserRole, float(peso_by_articulo.get(art_id, 0.0) or 0.0))
        self.table.blockSignals(False)
        self._reload_history()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() != 5:
            return
        row = item.row()
        teorico_txt = str(self.table.item(row, 4).text() if self.table.item(row, 4) else "0").replace(",", ".")
        conteo_txt = str(self.table.item(row, 5).text() if self.table.item(row, 5) else "").strip().replace(",", ".")
        teorico = float(teorico_txt or 0.0)
        conteo = float(conteo_txt or teorico)
        diff = conteo - teorico
        peso = float(self.table.item(row, 5).data(Qt.ItemDataRole.UserRole) or 0.0) if self.table.item(row, 5) else 0.0
        kg = diff * peso
        self.table.blockSignals(True)
        self.table.setItem(row, 6, QTableWidgetItem(f"{diff:.2f}"))
        self.table.setItem(row, 7, QTableWidgetItem(f"{kg:.2f} kg"))
        self.table.item(row, 6).setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.item(row, 7).setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.item(row, 6).setFlags(self.table.item(row, 6).flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.item(row, 7).setFlags(self.table.item(row, 7).flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.blockSignals(False)

    def _apply_adjustments(self) -> None:
        if not self._pending_ajustes:
            QMessageBox.information(self, "Inventarios", "No hay ajustes preparados. Pulsa 'Preparar ajustes'.")
            return
        aprobador = str(self.approver_input.text() or "").strip()
        contador = str(self.counter_input.text() or "").strip()
        if not contador or not aprobador:
            QMessageBox.warning(self, "Inventarios", "Indica Contador y Aprobador antes de aplicar.")
            return
        answer = QMessageBox.question(
            self,
            "Confirmar aprobación",
            f"Se aprobarán y aplicarán {len(self._pending_ajustes)} ajustes.\nContador: {contador}\nAprobador: {aprobador}\n¿Continuar?",
        )
        if answer != QMessageBox.StandardButton.Yes:
            return
        applied = self.inventory_service.apply_adjustments(
            pending=self._pending_ajustes,
            almacen_id=self._almacen_id,
            contador=contador,
            aprobador=aprobador,
        )
        self._pending_ajustes = []
        self.pending_label.setText(self.view.pending_label)
        QMessageBox.information(self, "Inventarios", f"Ajustes aplicados: {applied}")
        self.reload()

    def _prepare_adjustments(self) -> None:
        rows: list[dict[str, Any]] = []
        for i in range(self.table.rowCount()):
            conteo_cell = self.table.item(i, 5)
            if conteo_cell is None:
                continue
            key_item = self.table.item(i, 0)
            art_id, lote, cad_iso = key_item.data(Qt.ItemDataRole.UserRole) if key_item else ("", "", "")
            rows.append(
                {
                    "articulo_id": str(art_id or "").strip(),
                    "articulo_lote": str(lote or "").strip(),
                    "articulo_caducidad_iso": str(cad_iso or "").strip(),
                    "conteo_text": str(conteo_cell.text() or ""),
                    "teorico_text": str(self.table.item(i, 4).text() if self.table.item(i, 4) else "0"),
                    "peso": float(conteo_cell.data(Qt.ItemDataRole.UserRole) or 0.0),
                }
            )
        result = self.inventory_adjustment_preparation_service.prepare_adjustments(
            rows,
            almacen_id=self._almacen_id,
        )
        if result.status == "error":
            QMessageBox.warning(self, "Inventarios", result.message or "No se pudieron preparar los ajustes.")
            return
        if result.status == "no_differences":
            QMessageBox.information(self, "Inventarios", result.message)
            return
        self._pending_ajustes = result.pending
        self.pending_label.setText(f"Pendientes: {result.prepared_count}")
        QMessageBox.information(self, "Inventarios", result.message)

    def _reload_history(self) -> None:
        rows = self.inventory_service.history(self._almacen_id)
        self.history_table.setRowCount(len(rows))
        for row_idx, inv in enumerate(rows):
            values = [
                str(getattr(inv, "inventario_id", "") or "").strip(),
                getattr(inv, "fecha").strftime("%d/%m/%Y") if getattr(inv, "fecha", None) else "",
                str(getattr(inv, "contador", "") or "").strip(),
                str(getattr(inv, "aprobador", "") or "").strip(),
                str(int(getattr(inv, "lineas", 0) or 0)),
                str(int(getattr(inv, "ajustes", 0) or 0)),
                str(getattr(inv, "estado", "") or "").strip(),
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx == 0:
                    item.setData(Qt.ItemDataRole.UserRole, str(getattr(inv, "inventario_id", "") or "").strip())
                self.history_table.setItem(row_idx, col_idx, item)
        if rows:
            self.history_table.selectRow(0)
            self._reload_history_detail()
        else:
            self.history_detail_table.setRowCount(0)

    def _reload_history_detail(self) -> None:
        selected = self.history_table.selectionModel().selectedRows() if self.history_table.selectionModel() else []
        if not selected:
            self.history_detail_table.setRowCount(0)
            return
        row = selected[0].row()
        code_item = self.history_table.item(row, 0)
        inventario_id = str(code_item.data(Qt.ItemDataRole.UserRole) or "").strip() if code_item else ""
        if not inventario_id:
            self.history_detail_table.setRowCount(0)
            return
        payload = self.inventory_service.history_detail(inventario_id)
        detail_rows = payload.details
        items = payload.items
        ref_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip() for x in items}
        nombre_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip() for x in items}

        self.history_detail_table.setRowCount(len(detail_rows))
        for row_idx, det in enumerate(detail_rows):
            art_id = str(getattr(det, "articulo_id", "") or "").strip()
            cad = getattr(det, "articulo_caducidad", None)
            values = [
                ref_by_articulo.get(art_id, "") or art_id,
                nombre_by_articulo.get(art_id, "") or art_id,
                str(getattr(det, "articulo_lote", "") or "").strip(),
                cad.strftime("%d/%m/%Y") if cad else "",
                f"{float(getattr(det, 'teorico_uds', 0.0) or 0.0):.2f}",
                f"{float(getattr(det, 'conteo_uds', 0.0) or 0.0):.2f}",
                f"{float(getattr(det, 'diferencia_uds', 0.0) or 0.0):.2f}",
                f"{float(getattr(det, 'kg_ajuste', 0.0) or 0.0):.2f} kg",
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx in (4, 5, 6, 7):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.history_detail_table.setItem(row_idx, col_idx, item)

    def _selected_history_id(self) -> str:
        selected = self.history_table.selectionModel().selectedRows() if self.history_table.selectionModel() else []
        if not selected:
            return ""
        row = selected[0].row()
        code_item = self.history_table.item(row, 0)
        return str(code_item.data(Qt.ItemDataRole.UserRole) or "").strip() if code_item else ""

    def _export_history_excel(self) -> None:
        selected_id = self._selected_history_id()
        default_name = (
            f"inventario_historial_{selected_id[:8]}.xlsx"
            if selected_id
            else f"inventario_historial_{date.today().strftime('%Y%m%d')}.xlsx"
        )
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar historial inventarios", default_name, "Excel (*.xlsx)")
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        payload = self.inventory_service.export_payload(almacen_id=self._almacen_id, selected_id=selected_id)
        headers = payload.headers
        details = payload.details
        items = payload.items

        if not headers:
            QMessageBox.information(self, "Inventarios", "No hay historial para exportar con el filtro actual.")
            return

        ref_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_referencia_corta", "") or "").strip() for x in items}
        nombre_by_articulo = {str(getattr(x, "articulo_id", "") or "").strip(): str(getattr(x, "articulo_descripcion", "") or "").strip() for x in items}

        wb = Workbook()
        ws_head = wb.active
        ws_head.title = "Inventarios"
        ws_head.append(["inventario_id", "fecha", "almacen_id", "contador", "aprobador", "lineas", "ajustes", "estado"])
        for inv in headers:
            ws_head.append(
                [
                    str(getattr(inv, "inventario_id", "") or "").strip(),
                    getattr(inv, "fecha").isoformat() if getattr(inv, "fecha", None) else "",
                    str(getattr(inv, "almacen_id", "") or "").strip(),
                    str(getattr(inv, "contador", "") or "").strip(),
                    str(getattr(inv, "aprobador", "") or "").strip(),
                    int(getattr(inv, "lineas", 0) or 0),
                    int(getattr(inv, "ajustes", 0) or 0),
                    str(getattr(inv, "estado", "") or "").strip(),
                ]
            )

        ws_det = wb.create_sheet("Detalle")
        ws_det.append(
            [
                "inventario_id",
                "articulo_id",
                "ref",
                "nombre",
                "lote",
                "caducidad",
                "teorico_uds",
                "conteo_uds",
                "diferencia_uds",
                "kg_ajuste",
            ]
        )
        for det in details:
            art_id = str(getattr(det, "articulo_id", "") or "").strip()
            cad = getattr(det, "articulo_caducidad", None)
            ws_det.append(
                [
                    str(getattr(det, "inventario_id", "") or "").strip(),
                    art_id,
                    ref_by_articulo.get(art_id, "") or art_id,
                    nombre_by_articulo.get(art_id, "") or art_id,
                    str(getattr(det, "articulo_lote", "") or "").strip(),
                    cad.isoformat() if cad else "",
                    float(getattr(det, "teorico_uds", 0.0) or 0.0),
                    float(getattr(det, "conteo_uds", 0.0) or 0.0),
                    float(getattr(det, "diferencia_uds", 0.0) or 0.0),
                    float(getattr(det, "kg_ajuste", 0.0) or 0.0),
                ]
            )

        wb.save(file_path)
        QMessageBox.information(
            self,
            "Inventarios",
            f"Historial exportado.\n{file_path}\n\nInventarios: {len(headers)}\nLíneas detalle: {len(details)}",
        )

    def _export_count_template(self) -> None:
        default_name = f"inventario_conteo_{date.today().strftime('%Y%m%d')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar plantilla inventario", default_name, "Excel (*.xlsx)")
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Conteo"
        headers = ["articulo_id", "ref", "nombre", "lote", "caducidad", "teorico_uds", "conteo_uds"]
        ws.append(headers)
        for i in range(self.table.rowCount()):
            key_item = self.table.item(i, 0)
            art_id, lote, cad_iso = key_item.data(Qt.ItemDataRole.UserRole) if key_item else ("", "", "")
            ws.append(
                [
                    str(art_id or ""),
                    str(self.table.item(i, 0).text() if self.table.item(i, 0) else ""),
                    str(self.table.item(i, 1).text() if self.table.item(i, 1) else ""),
                    str(lote or ""),
                    cad_iso,
                    str(self.table.item(i, 4).text() if self.table.item(i, 4) else "0"),
                    "",
                ]
            )
        wb.save(file_path)
        QMessageBox.information(self, "Inventarios", f"Plantilla exportada:\n{file_path}")

    def _import_count_template(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar conteo inventario", "", "Excel (*.xlsx *.xlsm)")
        if not file_path:
            return
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(self, "Inventarios", "El archivo no contiene datos.")
            return
        header = [str(x or "").strip().lower() for x in rows[0]]
        idx_id, idx_lote, idx_conteo = _count_template_column_indexes(header)
        if idx_id < 0 or idx_lote < 0 or idx_conteo < 0:
            QMessageBox.warning(self, "Inventarios", "Faltan columnas requeridas: articulo_id, lote, conteo_uds.")
            return

        mapping = _count_template_mapping(rows[1:], idx_id, idx_lote, idx_conteo)

        loaded = 0
        self.table.blockSignals(True)
        for i in range(self.table.rowCount()):
            key_item = self.table.item(i, 0)
            art_id, lote, _cad_iso = key_item.data(Qt.ItemDataRole.UserRole) if key_item else ("", "", "")
            key = (str(art_id or "").strip(), str(lote or "").strip())
            if key not in mapping:
                continue
            self.table.item(i, 5).setText(str(mapping[key]))
            loaded += 1
        self.table.blockSignals(False)
        for i in range(self.table.rowCount()):
            cell = self.table.item(i, 5)
            if cell and str(cell.text() or "").strip():
                self._on_item_changed(cell)
        QMessageBox.information(self, "Inventarios", f"Conteos cargados: {loaded}")


class AnnualMonthlyOrdersTab(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._almacen_id = ""
        self._loading_filters = False
        self.service = MonthlyOrdersService()
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        filters = QHBoxLayout()
        filters.addWidget(QLabel("Año"))
        self.year_filter = QComboBox()
        self.year_filter.setMinimumWidth(95)
        self.year_filter.currentIndexChanged.connect(self.reload)
        filters.addWidget(self.year_filter)
        filters.addWidget(QLabel("Producto"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Ref. o nombre...")
        self.search_input.setMinimumWidth(320)
        self.search_input.textChanged.connect(self.reload)
        filters.addWidget(self.search_input, 1)
        refresh_btn = QPushButton("Refrescar")
        refresh_btn.setProperty("btnRole", "secondary")
        refresh_btn.clicked.connect(self.reload)
        filters.addWidget(refresh_btn)
        layout.addLayout(filters)

        self.table = QTableWidget(0, 18)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            """
            QTableWidget {
                font-size: 11px;
            }
            QTableWidget::item {
                padding: 2px 3px;
            }
            QTableWidget::item:focus {
                border: none;
                outline: 0;
            }
            QHeaderView::section {
                font-size: 11px;
                padding: 3px 2px;
            }
            """
        )
        header = self.table.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for col in range(2, 18):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)
        self.table.setHorizontalHeaderLabels(
            [
                "Ref.",
                "Producto",
                "Ene",
                "Feb",
                "Mar",
                "Abr",
                "May",
                "Jun",
                "Jul",
                "Ago",
                "Sep",
                "Oct",
                "Nov",
                "Dic",
                "Total",
                "Kg",
                "Ped.",
                "Ultimo",
            ]
        )
        self.table.setColumnWidth(0, 78)
        for col in range(2, 14):
            self.table.setColumnWidth(col, 54)
        self.table.setColumnWidth(14, 68)
        self.table.setColumnWidth(15, 92)
        self.table.setColumnWidth(16, 58)
        self.table.setColumnWidth(17, 82)
        self.table.itemDoubleClicked.connect(self._open_article_orders_dialog)
        layout.addWidget(self.table, 1)

    def set_almacen_filter(self, almacen_id: str) -> None:
        self._almacen_id = str(almacen_id or "").strip()
        self.reload()

    def _selected_year(self) -> int:
        return int(str(self.year_filter.currentData() or "0") or "0")

    def _reload_years(self, years: list[int]) -> None:
        current = self._selected_year()
        if current <= 0 and years:
            current = date.today().year if date.today().year in years else years[0]
        self._loading_filters = True
        try:
            self.year_filter.blockSignals(True)
            self.year_filter.clear()
            for year in years:
                self.year_filter.addItem(str(year), str(year))
            idx = self.year_filter.findData(str(current))
            self.year_filter.setCurrentIndex(idx if idx >= 0 else 0)
            self.year_filter.blockSignals(False)
        finally:
            self._loading_filters = False

    def reload(self) -> None:
        if self._loading_filters:
            return
        years = self.service.available_years_for(almacen_id=self._almacen_id)
        self._reload_years(years)
        rows = self.service.annual_product_matrix_for(
            year=self._selected_year(),
            almacen_id=self._almacen_id,
            search=self.search_input.text(),
        )
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            ref_item = QTableWidgetItem(row.referencia)
            ref_item.setData(Qt.ItemDataRole.UserRole, row.articulo_id)
            self.table.setItem(row_idx, 0, ref_item)
            self.table.setItem(row_idx, 1, QTableWidgetItem(row.descripcion))

            for month_idx, value in enumerate(row.monthly_quantities, start=2):
                item = SortableTableWidgetItem(f"{value:.2f}" if abs(value) > 1e-12 else "", value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row_idx, month_idx, item)

            total_item = SortableTableWidgetItem(f"{row.total_quantity:.2f}", row.total_quantity)
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            total_font = total_item.font()
            total_font.setBold(True)
            total_item.setFont(total_font)
            self.table.setItem(row_idx, 14, total_item)

            kg_item = SortableTableWidgetItem(f"{row.total_kg:.2f} kg", row.total_kg)
            kg_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            kg_font = kg_item.font()
            kg_font.setBold(True)
            kg_item.setFont(kg_font)
            self.table.setItem(row_idx, 15, kg_item)

            order_item = SortableTableWidgetItem(str(row.order_count), row.order_count)
            order_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 16, order_item)

            last_sort = row.last_order_date.toordinal() if row.last_order_date else 0
            last_item = SortableTableWidgetItem(
                row.last_order_date.strftime("%d/%m/%Y") if row.last_order_date else "",
                last_sort,
            )
            last_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 17, last_item)
        self.table.setSortingEnabled(True)

    def _open_article_orders_dialog(self, item: QTableWidgetItem) -> None:
        articulo_id = ""
        ref_item = self.table.item(item.row(), 0) if item is not None else None
        if ref_item is not None:
            articulo_id = str(ref_item.data(Qt.ItemDataRole.UserRole) or "").strip()
        if not articulo_id:
            return

        product_name = str(self.table.item(item.row(), 1).text() if self.table.item(item.row(), 1) else "").strip()
        product_ref = str(self.table.item(item.row(), 0).text() if self.table.item(item.row(), 0) else "").strip()
        rows = self.service.product_order_details_for(
            articulo_id=articulo_id,
            almacen_id=self._almacen_id,
        )

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Pedidos - {product_ref}".strip())
        dialog.resize(760, 520)
        layout = QVBoxLayout(dialog)
        title = QLabel(f"{product_ref} - {product_name}".strip(" -"))
        title.setProperty("role", "sectionTitle")
        layout.addWidget(title)

        table = QTableWidget(0, 5)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)
        table.setHorizontalHeaderLabels(["Fecha", "Pedido Nº", "Albarán", "Cantidad", "Kg"])
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        table.setColumnWidth(0, 105)
        table.setColumnWidth(1, 130)
        table.setColumnWidth(3, 95)
        table.setColumnWidth(4, 95)
        table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            values = [
                row.fecha.strftime("%d/%m/%Y") if row.fecha else "",
                row.pedido_numero,
                row.pedido_albaran_numero,
                f"{row.quantity:.2f}",
                f"{row.kg:.2f} kg",
            ]
            sort_values = [
                row.fecha.toordinal() if row.fecha else 0,
                row.pedido_numero,
                row.pedido_albaran_numero,
                row.quantity,
                row.kg,
            ]
            for col_idx, value in enumerate(values):
                cell = SortableTableWidgetItem(value, sort_values[col_idx])
                if col_idx in (3, 4):
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(row_idx, col_idx, cell)
        table.setSortingEnabled(True)
        layout.addWidget(table, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.exec()


class WarehousePage(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[Cliente] = []
        self.articles_tab: IngredientsIreksPage | None = None
        self.entradas_tab: MovimientosTab | None = None
        self.salidas_tab: MovimientosTab | None = None
        self.stock_tab: StockTab | None = None
        self.inventarios_tab: InventariosTab | None = None
        self.caducidad_tab: CaducidadTab | None = None
        self.monthly_orders_tab: AnnualMonthlyOrdersTab | None = None
        self.catalog_service = WarehouseCatalogService()
        self._build_ui()
        self.reload()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        title = QLabel("Almacen")
        title.setProperty("role", "pageTitle")
        layout.addWidget(title)

        row = QHBoxLayout()
        row.addWidget(QLabel("Cliente/Distribuidor"))
        self.almacen_combo = QComboBox()
        row.addWidget(self.almacen_combo, 2)
        self.almacen_combo.currentIndexChanged.connect(self._on_combo_filter_changed)

        refresh_btn = QPushButton("Refrescar")
        refresh_btn.setProperty("btnRole", "secondary")
        refresh_btn.clicked.connect(self.reload)
        row.addWidget(refresh_btn)
        row.addStretch(1)
        layout.addLayout(row)

        self.main_tabs = QTabWidget()
        self.articles_tab = IngredientsIreksPage(
            show_header=False,
            show_actions_ribbon=False,
            compact_mode=True,
            vm=IngredientWarehouseViewModel(),
        )
        self.entradas_tab = MovimientosTab(mode="in")
        self.salidas_tab = MovimientosTab(mode="out")
        self.stock_tab = StockTab()
        self.monthly_orders_tab = AnnualMonthlyOrdersTab()
        self.inventarios_tab = InventariosTab()
        self.caducidad_tab = CaducidadTab()
        self.main_tabs.addTab(self.articles_tab, "Artículos")
        self.main_tabs.addTab(self.entradas_tab, "Entradas")
        self.main_tabs.addTab(self.salidas_tab, "Salidas")
        self.main_tabs.addTab(self.stock_tab, "Stock")
        self.main_tabs.addTab(self.monthly_orders_tab, "Pedidos mensual")
        self.main_tabs.addTab(self.inventarios_tab, "Inventarios")
        self.main_tabs.addTab(self.caducidad_tab, "Caducidad")
        separator_idx = self.main_tabs.addTab(QWidget(), "|")
        self.main_tabs.setTabEnabled(separator_idx, False)
        self.main_tabs.addTab(self._build_fabricantes_tab(), "Fabricantes")
        self.main_tabs.addTab(OtrasReferenciasTab(), "Otras ref.")
        self.main_tabs.addTab(self._build_familias_tab(), "Familias")
        self.main_tabs.addTab(self._build_subfamilias_tab(), "Subfamilias")
        self.main_tabs.addTab(self._build_envases_tab(), "Envases")
        self._style_main_tabs()
        if self.entradas_tab is not None:
            self.entradas_tab.table.itemDoubleClicked.connect(self._open_article_from_entradas_row)
            self.entradas_tab.table.cellDoubleClicked.connect(self._open_article_from_entradas_cell)
        if self.salidas_tab is not None:
            self.salidas_tab.table.itemDoubleClicked.connect(self._open_article_from_salidas_row)
            self.salidas_tab.table.cellDoubleClicked.connect(self._open_article_from_salidas_cell)
        layout.addWidget(self.main_tabs, 1)

    def _style_main_tabs(self) -> None:
        bar = self.main_tabs.tabBar()
        self.main_tabs.setStyleSheet(
            "QTabBar::tab { padding: 8px 14px; }"
            "QTabBar::tab:!enabled { color: #A7B3C5; }"
        )
        for idx in range(self.main_tabs.count()):
            if idx in (0, 1, 2, 3, 4):
                bar.setTabTextColor(idx, QColor("#1E4FA1"))
            elif idx == 5:
                bar.setTabTextColor(idx, QColor("#A7B3C5"))
            else:
                bar.setTabTextColor(idx, QColor("#5E6C84"))

    def _build_placeholder_tab(self, name: str) -> QWidget:
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(8, 8, 8, 8)
        tab_layout.addWidget(QLabel(f"Pestaña {name} pendiente de implementar."), 1)
        return tab

    def _build_fabricantes_tab(self) -> QWidget:
        schema = [
            {"name": "fabricante_id", "label": "Fabricante_ID"},
            {"name": "fabricante_codigo", "label": "Fabricante_Codigo"},
            {"name": "fabricante_nombre", "label": "Fabricante_Nombre"},
        ]
        return EntityPage(
            title="Fabricantes",
            columns=[("fabricante_codigo", "Codigo"), ("fabricante_nombre", "Nombre")],
            schema=schema,
            list_fn=self._list_fabricantes,
            create_fn=self._create_fabricante,
            update_fn=self._update_fabricante,
            delete_fn=self._delete_fabricante,
            include_filters=False,
            import_fn=self._import_fabricantes,
            id_attr="fabricante_id",
        )

    def _build_familias_tab(self) -> QWidget:
        schema = [
            {"name": "fabricante_id", "label": "Fabricante_ID"},
            {"name": "articulo_familia_id", "label": "Articulo_Familia_ID"},
            {"name": "articulo_familia_nombre", "label": "Articulo_Familia_Nombre"},
            {"name": "articulo_familia_codigo", "label": "Articulo_Familia_Codigo"},
        ]
        return EntityPage(
            title="Familias",
            columns=[("articulo_familia_codigo", "Codigo"), ("articulo_familia_nombre", "Nombre")],
            schema=schema,
            list_fn=self._list_familias,
            create_fn=self._create_familia,
            update_fn=self._update_familia,
            delete_fn=self._delete_familia,
            include_filters=False,
            import_fn=self._import_familias,
            id_attr="articulo_familia_id",
        )

    def _build_subfamilias_tab(self) -> QWidget:
        schema = [
            {"name": "articulo_familia_id", "label": "Articulo_Familia_ID"},
            {"name": "articulo_subfamilia_id", "label": "Articulo_SubFamilia_ID"},
            {"name": "articulo_subfamilia_nombre", "label": "Articulo_SubFamilia_Nombre"},
            {"name": "articulo_subfamilia_codigo", "label": "Articulo_SubFamilia_Codigo"},
        ]
        return EntityPage(
            title="Subfamilias",
            columns=[("articulo_subfamilia_codigo", "Codigo"), ("articulo_subfamilia_nombre", "Nombre")],
            schema=schema,
            list_fn=self._list_subfamilias,
            create_fn=self._create_subfamilia,
            update_fn=self._update_subfamilia,
            delete_fn=self._delete_subfamilia,
            include_filters=False,
            import_fn=self._import_subfamilias,
            id_attr="articulo_subfamilia_id",
        )

    def _build_envases_tab(self) -> QWidget:
        schema = [
            {"name": "envase_id", "label": "Envase_ID"},
            {"name": "envase_codigo", "label": "Envase_Codigo"},
            {"name": "envase_nombre", "label": "Envase_Nombre"},
        ]
        return EntityPage(
            title="Envases",
            columns=[("envase_codigo", "Codigo"), ("envase_nombre", "Nombre")],
            schema=schema,
            list_fn=self._list_envases,
            create_fn=self._create_envase,
            update_fn=self._update_envase,
            delete_fn=self._delete_envase,
            include_filters=False,
            import_fn=self._import_envases,
            id_attr="envase_id",
        )

    def reload(self) -> None:
        self.rows = self.catalog_service.list_warehouse_clients()
        self.almacen_combo.blockSignals(True)
        self.almacen_combo.clear()
        self.almacen_combo.addItem("Todos", "")
        for item in self.rows:
            tipo = str(getattr(item, "cliente_tipo", "") or "").strip().lower()
            if tipo not in {"distribuidor", "directo", "cliente directo", "cliente_directo"}:
                continue
            label = str(item.cliente_nombre_comercial or "").strip() or str(item.cliente_nombre_fiscal or "").strip()
            if not label:
                continue
            cliente_id = str(item.cliente_id or "").strip()
            if not cliente_id:
                continue
            self.almacen_combo.addItem(label, cliente_id)
        self.almacen_combo.blockSignals(False)
        if self.almacen_combo.count() > 0:
            self.almacen_combo.setCurrentIndex(0)
        if self.articles_tab is not None:
            self.articles_tab.set_external_distributor_filter("")
        if self.entradas_tab is not None:
            self.entradas_tab.set_almacen_filter("")
        if self.salidas_tab is not None:
            self.salidas_tab.set_almacen_filter("")
        if self.stock_tab is not None:
            self.stock_tab.set_almacen_filter("")
        if self.monthly_orders_tab is not None:
            self.monthly_orders_tab.set_almacen_filter("")
        if self.inventarios_tab is not None:
            self.inventarios_tab.set_almacen_filter("")
        if self.caducidad_tab is not None:
            self.caducidad_tab.set_almacen_filter("")

    def _on_combo_filter_changed(self) -> None:
        if self.articles_tab is None:
            return
        selected = str(self.almacen_combo.currentData() or "")
        self.articles_tab.set_external_distributor_filter(selected)
        if self.entradas_tab is not None:
            self.entradas_tab.set_almacen_filter(selected)
        if self.salidas_tab is not None:
            self.salidas_tab.set_almacen_filter(selected)
        if self.stock_tab is not None:
            self.stock_tab.set_almacen_filter(selected)
        if self.monthly_orders_tab is not None:
            self.monthly_orders_tab.set_almacen_filter(selected)
        if self.inventarios_tab is not None:
            self.inventarios_tab.set_almacen_filter(selected)
        if self.caducidad_tab is not None:
            self.caducidad_tab.set_almacen_filter(selected)

    def _open_article_from_salidas_row(self, item: QTableWidgetItem) -> None:
        if self.salidas_tab is None or self.articles_tab is None:
            return
        articulo_id = self.salidas_tab.selected_articulo_id_from_row(item.row())
        if not articulo_id:
            return
        self.main_tabs.setCurrentWidget(self.articles_tab)
        focus_fn = getattr(self.articles_tab, "focus_article_and_open_salidas", None)
        if callable(focus_fn):
            focus_fn(articulo_id)

    def _open_article_from_salidas_cell(self, row: int, _column: int) -> None:
        if self.salidas_tab is None or self.articles_tab is None:
            return
        articulo_id = self.salidas_tab.selected_articulo_id_from_row(row)
        if not articulo_id:
            return
        self.main_tabs.setCurrentWidget(self.articles_tab)
        focus_fn = getattr(self.articles_tab, "focus_article_and_open_salidas", None)
        if callable(focus_fn):
            focus_fn(articulo_id)

    def _open_article_from_entradas_row(self, item: QTableWidgetItem) -> None:
        if self.entradas_tab is None or self.articles_tab is None:
            return
        articulo_id = self.entradas_tab.selected_articulo_id_from_row(item.row())
        if not articulo_id:
            return
        self.main_tabs.setCurrentWidget(self.articles_tab)
        focus_fn = getattr(self.articles_tab, "focus_article_and_open_entradas", None)
        if callable(focus_fn):
            focus_fn(articulo_id)

    def _open_article_from_entradas_cell(self, row: int, _column: int) -> None:
        if self.entradas_tab is None or self.articles_tab is None:
            return
        articulo_id = self.entradas_tab.selected_articulo_id_from_row(row)
        if not articulo_id:
            return
        self.main_tabs.setCurrentWidget(self.articles_tab)
        focus_fn = getattr(self.articles_tab, "focus_article_and_open_entradas", None)
        if callable(focus_fn):
            focus_fn(articulo_id)

    def _list_fabricantes(self, term: str) -> list[Fabricante]:
        return self.catalog_service.list_fabricantes(term)

    def _create_fabricante(self, payload: dict) -> None:
        self.catalog_service.create_fabricante(payload)

    def _update_fabricante(self, fabricante_id: str, payload: dict) -> None:
        self.catalog_service.update_fabricante(fabricante_id, payload)

    def _delete_fabricante(self, fabricante_id: str) -> bool:
        return self.catalog_service.delete_fabricante(fabricante_id)

    def _import_fabricantes(self, file_path: str) -> tuple[int, list[str]]:
        return self.catalog_service.import_fabricantes(file_path)

    def _list_familias(self, term: str) -> list[Familia]:
        return self.catalog_service.list_familias(term)

    def _create_familia(self, payload: dict) -> None:
        self.catalog_service.create_familia(payload)

    def _update_familia(self, familia_id: str, payload: dict) -> None:
        self.catalog_service.update_familia(familia_id, payload)

    def _delete_familia(self, familia_id: str) -> bool:
        return self.catalog_service.delete_familia(familia_id)

    def _import_familias(self, file_path: str) -> tuple[int, list[str]]:
        return self.catalog_service.import_familias(file_path)

    def _list_subfamilias(self, term: str) -> list[Subfamilia]:
        return self.catalog_service.list_subfamilias(term)

    def _create_subfamilia(self, payload: dict) -> None:
        self.catalog_service.create_subfamilia(payload)

    def _update_subfamilia(self, subfamilia_id: str, payload: dict) -> None:
        self.catalog_service.update_subfamilia(subfamilia_id, payload)

    def _delete_subfamilia(self, subfamilia_id: str) -> bool:
        return self.catalog_service.delete_subfamilia(subfamilia_id)

    def _import_subfamilias(self, file_path: str) -> tuple[int, list[str]]:
        return self.catalog_service.import_subfamilias(file_path)

    def _list_envases(self, term: str) -> list[Envase]:
        return self.catalog_service.list_envases(term)

    def _create_envase(self, payload: dict) -> None:
        self.catalog_service.create_envase(payload)

    def _update_envase(self, envase_id: str, payload: dict) -> None:
        self.catalog_service.update_envase(envase_id, payload)

    def _delete_envase(self, envase_id: str) -> bool:
        return self.catalog_service.delete_envase(envase_id)

    def _import_envases(self, file_path: str) -> tuple[int, list[str]]:
        return self.catalog_service.import_envases(file_path)
