from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import warnings
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, col, select

from app.core.database import engine
from app.models import Cliente, Distribuidor, IngredienteIreks, ReferenciaDistribuidor, VentaImportLote, VentaMensualRaw


class IgsaSalesWorkbookFlowService:
    def __init__(self) -> None:
        self._igsa_unit_codes = {"LPAO", "LPAOP", "555", "777"}

    def parse_igsa_workbook_by_sheets(self, file_path: Path) -> tuple[list[object], list[str]]:
        if not file_path.exists():
            return [], [f"No existe el archivo: {file_path}"]
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style, apply openpyxl's default",
                category=UserWarning,
            )
            workbook = load_workbook(file_path, data_only=True)
        sheetnames = list(workbook.sheetnames or [])
        if not sheetnames:
            return [], ["El libro no contiene hojas."]

        rows_out: list[object] = []
        errors: list[str] = []
        for sheet_idx, sheet_name in enumerate(sheetnames, start=1):
            ws_obj = workbook[sheet_name]
            ws = ws_obj if isinstance(ws_obj, Worksheet) else None
            if ws is None:
                continue
            if sheet_idx == 1:
                tipo = "venta"
            elif sheet_idx == 2:
                tipo = "promocion"
            else:
                tipo = "muestras"
            es_sc = self._is_sc_tipo(tipo)
            periodo = self._infer_period(ws, file_path)
            if not periodo:
                errors.append(f"{sheet_name}: no se pudo inferir periodo (AAAA-MM).")
                continue
            for row_idx in range(1, ws.max_row + 1):
                ref_distribuidor = str(ws.cell(row=row_idx, column=1).value or "").strip()
                codigo_raw = self._normalize_code(ws.cell(row=row_idx, column=2).value)
                codigo_compact = re.sub(r"\s+", "", str(codigo_raw or "").upper())
                if codigo_compact in {"LPAO", "LPAOP"}:
                    codigo = codigo_compact
                else:
                    code_match = re.match(r"^(D)?(\d+)", codigo_compact)
                    codigo = f"{code_match.group(1) or ''}{code_match.group(2)}" if code_match else codigo_compact
                descripcion = str(ws.cell(row=row_idx, column=3).value or "").strip()
                peso_envase = self._to_float(ws.cell(row=row_idx, column=4).value)
                cantidad_total = self._to_float(ws.cell(row=row_idx, column=5).value)
                if cantidad_total <= 0:
                    continue
                if codigo and codigo not in {"LPAO", "LPAOP"} and not re.fullmatch(r"(?:D\d+|\d+)", codigo):
                    errors.append(
                        f"{sheet_name} fila {row_idx}: referencia fabricante invalida '{codigo}' "
                        "(solo se admite numerico o prefijo D al inicio)."
                    )
                    continue
                if not codigo or not descripcion:
                    continue
                lotes: list[tuple[float, str, bool]] = []
                for col_idx in range(6, 14):
                    raw = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
                    if not raw:
                        continue
                    raw_norm = re.sub(r"\s+", "", raw)
                    match = re.match(r"(?:(\d+(?:[.,]\d+)?)\s*)?L\.?(.+)$", raw_norm, flags=re.IGNORECASE)
                    if not match:
                        if es_sc:
                            continue
                        lotes.append((1.0, raw_norm, True))
                        continue
                    qty_txt = str(match.group(1) or "").strip()
                    lote = str(match.group(2) or "").strip()
                    qty_missing = not bool(qty_txt)
                    qty = self._to_float(qty_txt) if qty_txt else 1.0
                    if qty <= 0:
                        errors.append(f"{sheet_name} fila {row_idx}: cantidad de lote no valida en columna {col_idx}.")
                        continue
                    if not lote:
                        errors.append(f"{sheet_name} fila {row_idx}: lote vacio en columna {col_idx}.")
                        continue
                    lotes.append((qty, lote, qty_missing))
                if len(lotes) == 1 and lotes[0][2] and cantidad_total > 0:
                    lotes = [(float(cantidad_total), lotes[0][1], False)]
                suma_lotes = sum(x[0] for x in lotes)
                if not lotes:
                    rows_out.append(
                        self._build_parsed_line(
                            sheet_name=sheet_name,
                            source_row=row_idx,
                            tipo=tipo,
                            periodo=periodo,
                            ref_distribuidor=ref_distribuidor,
                            codigo=codigo,
                            descripcion=descripcion,
                            peso_envase=peso_envase,
                            cantidad_total=cantidad_total,
                            cantidad_lote=float(cantidad_total),
                            lote="",
                            es_sc=es_sc,
                            suma_cantidad_lotes=float(cantidad_total),
                        )
                    )
                    continue
                if abs(suma_lotes - cantidad_total) > 0.001:
                    errors.append(
                        f"{sheet_name} fila {row_idx}: suma lotes={suma_lotes:.3f} distinta de E={cantidad_total:.3f}."
                    )
                for qty_lote, lote, _qty_missing in lotes:
                    rows_out.append(
                        self._build_parsed_line(
                            sheet_name=sheet_name,
                            source_row=row_idx,
                            tipo=tipo,
                            periodo=periodo,
                            ref_distribuidor=ref_distribuidor,
                            codigo=codigo,
                            descripcion=descripcion,
                            peso_envase=peso_envase,
                            cantidad_total=cantidad_total,
                            cantidad_lote=float(qty_lote),
                            lote=lote,
                            es_sc=es_sc,
                            suma_cantidad_lotes=float(suma_lotes),
                        )
                    )
        return rows_out, errors

    def build_igsa_workbook_preview(
        self,
        lines: list[object],
        cliente_id: str,
    ) -> tuple[list[dict[str, object]], list[str]]:
        if not lines:
            return [], []
        preview_rows: list[dict[str, object]] = []
        errors: list[str] = []
        with Session(engine) as session:
            distribuidor_ids = self._resolve_igsa_distribuidor_ids(session, cliente_id)
            ref_rows = (
                list(
                    session.exec(
                        select(ReferenciaDistribuidor).where(col(ReferenciaDistribuidor.distribuidor_id).in_(sorted(distribuidor_ids)))
                    )
                )
                if distribuidor_ids
                else []
            )
            articulo_ids = {str(getattr(x, "articulo_id", "") or "").strip() for x in ref_rows}
            articulo_ids = {x for x in articulo_ids if x}
            products = (
                list(session.exec(select(IngredienteIreks).where(col(IngredienteIreks.articulo_id).in_(sorted(articulo_ids)))))
                if articulo_ids
                else []
            )
            product_by_id = {str(getattr(x, "articulo_id", "") or "").strip(): x for x in products}
            ref_map: dict[str, str] = {}
            for ref in ref_rows:
                aid = str(getattr(ref, "articulo_id", "") or "").strip()
                raw = str(getattr(ref, "articulo_referencia_distribuidor", "") or "").strip()
                if not aid or not raw:
                    continue
                for cand in self._code_candidates(raw):
                    ref_map[cand] = aid

            for line in lines:
                ref_distribuidor = str(getattr(line, "ref_distribuidor", "") or "").strip()
                descripcion = str(getattr(line, "descripcion", "") or "").strip()
                cantidad_lote = float(getattr(line, "cantidad_lote", 0.0) or 0.0)
                periodo = str(getattr(line, "periodo", "") or "").strip()
                tipo = str(getattr(line, "tipo", "") or "").strip()
                lote = str(getattr(line, "lote", "") or "").strip()
                source_row = int(getattr(line, "source_row", 0) or 0)
                sheet_name = str(getattr(line, "sheet_name", "") or "").strip()
                aid = ""
                for cand in self._code_candidates(ref_distribuidor):
                    aid = ref_map.get(cand, "")
                    if aid:
                        break
                if not aid:
                    errors.append(f"{sheet_name} fila {source_row}: sin ficha por ref. distribuidor '{ref_distribuidor}'")
                    continue
                product = product_by_id.get(aid)
                if product is None:
                    errors.append(f"{sheet_name} fila {source_row}: articulo_id sin ficha '{aid}'")
                    continue
                peso_envase = float(getattr(product, "articulo_envase_peso_total", 0.0) or 0.0)
                if peso_envase <= 0:
                    peso_envase = float(getattr(product, "articulo_envase_peso", 0.0) or 0.0)
                ref_fabricante = str(getattr(product, "articulo_referencia_corta", "") or "").strip() or str(
                    getattr(product, "articulo_referencia", "") or ""
                ).strip()
                descripcion_final = str(getattr(product, "articulo_descripcion", "") or "").strip() or descripcion
                preview_rows.append(
                    {
                        "periodo": periodo,
                        "ref_distribuidor": ref_distribuidor,
                        "ref_fabricante": ref_fabricante,
                        "descripcion": descripcion_final,
                        "peso_envase": peso_envase,
                        "num_envases": cantidad_lote,
                        "tot_kg": cantidad_lote * peso_envase,
                        "lote": lote,
                        "tipo": tipo,
                        "articulo_id": aid,
                    }
                )
        return preview_rows, errors

    def import_igsa_workbook_lines(
        self,
        lines: list[object],
        cliente_id: str,
        *,
        force_reimport: bool = False,
        sync_warehouse_callback: Callable[[Session, list[VentaMensualRaw]], None] | None = None,
    ):
        from app.services.sales_reconciliation_service import SalesOpResult

        clean_cliente_id = str(cliente_id or "").strip()
        if not clean_cliente_id:
            return SalesOpResult(False, "Cliente/Distribuidor IGSA no valido.")
        if not lines:
            return SalesOpResult(False, "No hay lineas para importar.")
        periodos = sorted({str(getattr(x, "periodo", "") or "").strip() for x in lines if str(getattr(x, "periodo", "") or "").strip()})
        if not periodos:
            return SalesOpResult(False, "No se detecto periodo valido en las lineas.")

        payload_hash = hashlib.sha256(
            "|".join(
                sorted(
                    [
                        f"{getattr(x, 'periodo', '')}|{getattr(x, 'sheet_name', '')}|{getattr(x, 'source_row', 0)}|{getattr(x, 'codigo', '')}|{getattr(x, 'cantidad_lote', 0.0)}|{getattr(x, 'lote', '')}|{getattr(x, 'tipo', '')}"
                        for x in lines
                    ]
                )
            ).encode("utf-8")
        ).hexdigest()

        with Session(engine) as session:
            distribuidor_ids = self._resolve_igsa_distribuidor_ids(session, clean_cliente_id)
            existing = session.exec(
                select(VentaImportLote).where(
                    VentaImportLote.fuente == "igsa_book",
                    VentaImportLote.archivo_hash == payload_hash,
                )
            ).first()
            if existing is not None and not force_reimport:
                return SalesOpResult(True, "El libro ya estaba importado.", imported=0, incidencias=0)
            if existing is not None and force_reimport:
                old_rows = list(session.exec(select(VentaMensualRaw).where(VentaMensualRaw.lote_id == existing.lote_id)))
                for old_row in old_rows:
                    session.delete(old_row)
                session.delete(existing)

            for periodo in periodos:
                stmt = select(VentaMensualRaw).where(
                    col(VentaMensualRaw.fuente) == "igsa_book",
                    col(VentaMensualRaw.periodo) == periodo,
                )
                for old_row in session.exec(stmt):
                    session.delete(old_row)

            lote = VentaImportLote(
                lote_id=str(uuid4()),
                fuente="igsa_book",
                cliente_id=clean_cliente_id,
                periodo=periodos[0],
                archivo_nombre="IGSA workbook sheets",
                archivo_hash=payload_hash,
                estado="procesado",
            )
            session.add(lote)
            session.flush()

            ref_rows = (
                list(
                    session.exec(
                        select(ReferenciaDistribuidor).where(col(ReferenciaDistribuidor.distribuidor_id).in_(sorted(distribuidor_ids)))
                    )
                )
                if distribuidor_ids
                else []
            )
            articulo_ids = {str(getattr(x, "articulo_id", "") or "").strip() for x in ref_rows}
            articulo_ids = {x for x in articulo_ids if x}
            products = (
                list(session.exec(select(IngredienteIreks).where(col(IngredienteIreks.articulo_id).in_(sorted(articulo_ids)))))
                if articulo_ids
                else []
            )
            product_by_id = {str(getattr(x, "articulo_id", "") or "").strip(): x for x in products}
            ref_map: dict[str, str] = {}
            for ref in ref_rows:
                aid = str(getattr(ref, "articulo_id", "") or "").strip()
                raw = str(getattr(ref, "articulo_referencia_distribuidor", "") or "").strip()
                if not aid or not raw:
                    continue
                for cand in self._code_candidates(raw):
                    ref_map[cand] = aid

            skipped = 0
            skipped_details: list[str] = []
            inserted_rows: list[VentaMensualRaw] = []
            for line in lines:
                code = self._normalize_code(getattr(line, "codigo", ""))
                ref_dist = self._normalize_code(getattr(line, "ref_distribuidor", ""))
                articulo_id = ""
                for cand in self._code_candidates(ref_dist):
                    articulo_id = ref_map.get(cand, "")
                    if articulo_id:
                        break
                source_row = int(getattr(line, "source_row", 0) or 0)
                sheet_name = str(getattr(line, "sheet_name", "") or "").strip()
                if not articulo_id:
                    skipped += 1
                    skipped_details.append(
                        f"{sheet_name} fila {source_row} ({code}): sin mapeo por ref distribuidor '{getattr(line, 'ref_distribuidor', '')}'"
                    )
                    continue
                product = product_by_id.get(articulo_id)
                if product is None:
                    skipped += 1
                    skipped_details.append(
                        f"{sheet_name} fila {source_row} ({code}): articulo_id {articulo_id} sin ficha"
                    )
                    continue
                descripcion = str(getattr(product, "articulo_descripcion", "") or "").strip() or str(getattr(line, "descripcion", "") or "").strip()
                peso_master = float(getattr(product, "articulo_envase_peso_total", 0.0) or 0.0)
                if peso_master <= 0:
                    peso_master = float(getattr(product, "articulo_envase_peso", 0.0) or 0.0)
                peso_fuente = peso_master if peso_master > 0 else float(getattr(line, "peso_envase", 0.0) or 0.0)
                cantidad_lote = float(getattr(line, "cantidad_lote", 0.0) or 0.0)
                kilos = cantidad_lote * float(peso_fuente or 0.0)
                is_unit_article = code in self._igsa_unit_codes
                if kilos <= 0:
                    if is_unit_article and cantidad_lote > 0:
                        kilos = cantidad_lote
                    else:
                        skipped += 1
                        skipped_details.append(
                            f"{sheet_name} fila {source_row} ({code}): kilos no validos (cantidad={cantidad_lote}, peso_ficha={peso_master}, peso_excel={getattr(line, 'peso_envase', 0.0)})"
                        )
                        continue
                tipo_norm = self._normalize_igsa_tipo(getattr(line, "tipo", ""))
                venta_kilos = 0.0 if self._is_sc_tipo(tipo_norm) else kilos
                venta_kilos_sc = kilos if self._is_sc_tipo(tipo_norm) else 0.0
                payload = {
                    "anio": int(str(getattr(line, "periodo", "")).split("-")[0]) if "-" in str(getattr(line, "periodo", "")) else 0,
                    "mes": int(str(getattr(line, "periodo", "")).split("-")[1]) if "-" in str(getattr(line, "periodo", "")) else 0,
                    "tipo": tipo_norm,
                    "ref_distribuidor": str(getattr(line, "ref_distribuidor", "") or "").strip(),
                    "cantidad_lote": cantidad_lote,
                    "lote": str(getattr(line, "lote", "") or "").strip(),
                    "articulo_id": articulo_id,
                }
                row = VentaMensualRaw(
                    raw_id=str(uuid4()),
                    lote_id=str(lote.lote_id or "").strip(),
                    fuente="igsa_book",
                    cliente_id=clean_cliente_id,
                    periodo=str(getattr(line, "periodo", "") or "").strip(),
                    articulo_codigo_origen=code,
                    articulo_id=articulo_id,
                    articulo_descripcion_origen=descripcion,
                    venta_kilos=venta_kilos,
                    venta_kilos_sc=venta_kilos_sc,
                    venta_euros=0.0,
                    payload_json=json.dumps(payload, ensure_ascii=False),
                )
                session.add(row)
                inserted_rows.append(row)

            session.commit()
            if not inserted_rows:
                return SalesOpResult(False, "No se pudo importar ninguna fila valida.", imported=0, incidencias=skipped)
            if sync_warehouse_callback is not None:
                sync_warehouse_callback(session, inserted_rows)
            session.commit()
        msg = "Importacion IGSA (libro por hojas) completada."
        if skipped_details:
            preview = " | ".join(skipped_details[:6])
            extra = "" if len(skipped_details) <= 6 else f" | ... y {len(skipped_details) - 6} mas."
            msg += f"\nOmitidas: {preview}{extra}"
        return SalesOpResult(True, msg, imported=len(inserted_rows), incidencias=skipped)

    def _build_parsed_line(self, **kwargs: object):
        from app.services.sales_reconciliation_service import IgsaWorkbookParsedLine

        return IgsaWorkbookParsedLine(**kwargs)

    def _infer_period(self, ws: Worksheet, file_path: Path) -> str:
        month_map = {
            "enero": 1,
            "ene": 1,
            "febrero": 2,
            "feb": 2,
            "marzo": 3,
            "mar": 3,
            "abril": 4,
            "abr": 4,
            "mayo": 5,
            "junio": 6,
            "jun": 6,
            "julio": 7,
            "jul": 7,
            "agosto": 8,
            "ago": 8,
            "septiembre": 9,
            "setiembre": 9,
            "sep": 9,
            "set": 9,
            "octubre": 10,
            "oct": 10,
            "noviembre": 11,
            "nov": 11,
            "diciembre": 12,
            "dic": 12,
        }
        scan_lines: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), min_col=1, max_col=4, values_only=True):
            for cell in row:
                txt = str(cell or "").strip()
                if txt:
                    scan_lines.append(txt)
        joined = " ".join(scan_lines).lower()
        joined = unicodedata.normalize("NFD", joined)
        joined = "".join(ch for ch in joined if unicodedata.category(ch) != "Mn")
        year_match = re.search(r"(20\d{2})", joined)
        yy_match = re.search(r"\b(\d{2})\b", joined)
        month_num = 0
        for key, val in month_map.items():
            if re.search(rf"\b{re.escape(key)}\b", joined):
                month_num = val
                break
        year = 0
        if year_match:
            year = int(year_match.group(1))
        elif yy_match:
            two = int(yy_match.group(1))
            year = 2000 + two if two <= 79 else 1900 + two
        if year > 0 and 1 <= month_num <= 12:
            return f"{year:04d}-{month_num:02d}"
        stem = unicodedata.normalize("NFD", str(file_path.stem or "").lower())
        stem = "".join(ch for ch in stem if unicodedata.category(ch) != "Mn")
        yy_file = re.search(r"\b(\d{2})\b", stem)
        month_file = 0
        for key, val in month_map.items():
            if re.search(rf"\b{re.escape(key)}\b", stem):
                month_file = val
                break
        if yy_file and month_file:
            two = int(yy_file.group(1))
            year_file = 2000 + two if two <= 79 else 1900 + two
            return f"{year_file:04d}-{month_file:02d}"
        return ""

    def _code_candidates(self, code: object) -> list[str]:
        base = self._normalize_code(code)
        clean = re.sub(r"\s+", "", str(base or "").upper())
        if not clean:
            return []
        out = [clean]
        if clean.startswith("D") and clean[1:].isdigit():
            out.append(clean[1:])
        elif clean.isdigit():
            out.append(f"D{clean}")
        return list(dict.fromkeys(out))

    def _resolve_igsa_distribuidor_ids(self, session: Session, cliente_id: str) -> set[str]:
        ids: set[str] = set()
        clean_cliente_id = str(cliente_id or "").strip()
        if clean_cliente_id:
            ids.add(clean_cliente_id)
        cliente = session.get(Cliente, clean_cliente_id) if clean_cliente_id else None
        distribuidor_id = str(getattr(cliente, "distribuidor_id", "") or "").strip() if cliente else ""
        if distribuidor_id:
            ids.add(distribuidor_id)
        if not distribuidor_id:
            dist_rows = list(session.exec(select(Distribuidor)))
            for row in dist_rows:
                comercial = str(getattr(row, "distribuidor_nombre_comercial", "") or "").strip().lower()
                fiscal = str(getattr(row, "distribuidor_razon_social", "") or "").strip().lower()
                if "igsa" in comercial or "igsa" in fiscal:
                    ids.add(str(getattr(row, "distribuidor_id", "") or "").strip())
                    break
        return {x for x in ids if x}

    def _is_sc_tipo(self, tipo: str) -> bool:
        t = self._normalize_igsa_tipo(tipo)
        return t in {"s/c", "muestra", "muestras", "promocion", "promociones"}

    def _normalize_code(self, value) -> str:
        text = str(value or "").strip().upper()
        if not text:
            return ""
        if re.fullmatch(r"\d+(\.0+)?", text):
            return str(int(float(text)))
        return text

    def _normalize_search_text(self, value) -> str:
        text = str(value or "").strip().lower()
        normalized = unicodedata.normalize("NFD", text)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        return re.sub(r"\s+", " ", normalized)

    def _normalize_igsa_tipo(self, value) -> str:
        raw = str(value or "").strip().lower()
        if not raw:
            return ""
        normalized = unicodedata.normalize("NFD", raw)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        normalized = normalized.replace(" ", "")
        if normalized in {"venta", "ventas"}:
            return "venta"
        if normalized in {"s/c", "c/s", "sc"}:
            return "s/c"
        if normalized in {"muestra", "muestras"}:
            return "muestras"
        if normalized in {"promocion", "promociones"}:
            return "promociones"
        return normalized

    def _to_int(self, value) -> int:
        try:
            txt = str(value or "").strip()
            if not txt:
                return 0
            return int(float(txt.replace(",", ".")))
        except Exception:
            return 0

    def _to_float(self, value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            try:
                return float(value)
            except Exception:
                return 0.0
        text = str(value).strip()
        if not text:
            return 0.0
        match = re.search(r"[-+]?\d+(?:[.,]\d+)?", text)
        if match is None:
            return 0.0
        try:
            return float(match.group(0).replace(",", "."))
        except Exception:
            return 0.0
