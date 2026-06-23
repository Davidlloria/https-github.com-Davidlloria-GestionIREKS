from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date
import hashlib
import json
import math
from pathlib import Path
import re
import unicodedata
import warnings
from uuid import NAMESPACE_URL, uuid4, uuid5

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, col, select

from app.core.config import BASE_DIR
from app.core.database import engine
from app.models import AlmacenMovimiento, Cliente, Distribuidor, Fabricante, Familia, IngredienteIreks, ReferenciaDistribuidor, Subfamilia, VentaImportLote, VentaMensualRaw
from app.services.igsa_sales_pdf_flow_service import IgsaSalesPdfFlowService
from app.services.igsa_sales_workbook_flow_service import IgsaSalesWorkbookFlowService
from app.services.sales_annual_comparison_service import SalesAnnualComparisonService
from app.services.sales_annual_comparison_service import SalesComparisonRow


SALES_CLIENT_TYPES = {"distribuidor", "directo", "cliente directo", "cliente_directo"}


@dataclass
class SalesOpResult:
    ok: bool
    message: str
    imported: int = 0
    incidencias: int = 0


@dataclass
class IgsaPdfParsedLine:
    source_file: str
    doc_type: str
    fecha: str
    ref_pedido: str
    codigo: str
    descripcion: str
    kilos: float
    envases: float
    emb: float
    precio: float
    descuento_pct: float
    total: float
    iva: float
    lote: str
    carga: str
    cons_pref: str


@dataclass
class IgsaWorkbookParsedLine:
    sheet_name: str
    source_row: int
    tipo: str
    periodo: str
    ref_distribuidor: str
    codigo: str
    descripcion: str
    peso_envase: float
    cantidad_total: float
    cantidad_lote: float
    lote: str
    es_sc: bool
    suma_cantidad_lotes: float


class SalesReconciliationService:
    def __init__(self) -> None:
        self._igsa_unit_codes = {"LPAO", "LPAOP", "555", "777"}
        self._igsa_rows_cache: list[dict[str, object]] | None = None
        self._igsa_mtime_ns: int | None = None
        self._igsa_pdf_flow_service = IgsaSalesPdfFlowService()
        self._igsa_workbook_flow_service = IgsaSalesWorkbookFlowService()
        self._sales_annual_comparison_service = SalesAnnualComparisonService(db_engine=engine)

    def _is_sc_tipo(self, tipo: str) -> bool:
        t = self._normalize_igsa_tipo(tipo)
        return t in {"s/c", "muestra", "muestras", "promocion", "promociones"}

    def _resolve_igsa_distribuidor_ids(self, session: Session, cliente_id: str) -> set[str]:
        ids: set[str] = set()
        clean_cliente_id = str(cliente_id or "").strip()
        if clean_cliente_id:
            ids.add(clean_cliente_id)
        cliente = session.get(Cliente, clean_cliente_id) if clean_cliente_id else None
        distribuidor_id = str(getattr(cliente, "distribuidor_id", "") or "").strip() if cliente else ""
        if distribuidor_id:
            ids.add(distribuidor_id)
        if not distribuidor_id:
            dist_rows = list(session.exec(select(Distribuidor)))
            for row in dist_rows:
                comercial = str(getattr(row, "distribuidor_nombre_comercial", "") or "").strip().lower()
                fiscal = str(getattr(row, "distribuidor_razon_social", "") or "").strip().lower()
                if "igsa" in comercial or "igsa" in fiscal:
                    ids.add(str(getattr(row, "distribuidor_id", "") or "").strip())
                    break
        return {x for x in ids if x}

    def import_ireks_json(self, file_path: Path) -> SalesOpResult:
        try:
            data = self._read_json(file_path)
        except ValueError as exc:
            return SalesOpResult(False, str(exc))
        rows_data, default_year, default_month, default_cliente_id = self._normalize_ireks_json_payload(data)
        if not rows_data:
            return SalesOpResult(False, "El JSON de IREKS no contiene filas.")

        file_hash = self._file_hash(file_path)
        with Session(engine) as session:
            existing = session.exec(
                select(VentaImportLote).where(
                    VentaImportLote.fuente == "ireks",
                    VentaImportLote.archivo_hash == file_hash,
                )
            ).first()
            if existing is not None:
                return SalesOpResult(True, "El archivo ya estaba importado.", imported=0, incidencias=0)

            rows: list[VentaMensualRaw] = []
            skipped = 0
            first_period = ""
            first_cliente_id = ""
            client_rows = list(
                session.exec(
                    select(Cliente.cliente_id, Cliente.cliente_nombre_comercial, Cliente.cliente_nombre_fiscal, Cliente.cliente_tipo)
                )
            )
            allowed_by_id: dict[str, str] = {}
            allowed_by_name: dict[str, str] = {}
            for cid, ncom, nfis, ctype in client_rows:
                tipo = str(ctype or "").strip().lower()
                if tipo not in SALES_CLIENT_TYPES:
                    continue
                cliente_id_val = str(cid or "").strip()
                if not cliente_id_val:
                    continue
                allowed_by_id[cliente_id_val] = cliente_id_val
                for raw_name in (ncom, nfis):
                    norm_name = self._normalize_search_text(raw_name)
                    if norm_name:
                        allowed_by_name[norm_name] = cliente_id_val
            for item in rows_data:
                if not isinstance(item, dict):
                    skipped += 1
                    continue

                year = self._to_int(self._get_any(item, "venta_Anio", "venta_Anio".lower(), "anio", "año"))
                month = self._parse_month(self._get_any(item, "venta_Mes", "mes", "venta_Mes".lower()))
                if year <= 0:
                    year = default_year
                if month < 1 or month > 12:
                    month = default_month
                if year <= 0 or month < 1 or month > 12:
                    skipped += 1
                    continue

                periodo = f"{year:04d}-{month:02d}"
                if not first_period:
                    first_period = periodo
                cliente_id = str(self._get_any(item, "Cliente_ID", "cliente_id", "ClienteId", "clienteid") or "").strip()
                if not cliente_id:
                    cliente_id = default_cliente_id
                if cliente_id and not first_cliente_id:
                    first_cliente_id = cliente_id

                code = self._normalize_code(self._get_any(item, "Código", "CÃ³digo", "Codigo", "codigo"))
                articulo_id = str(self._get_any(item, "Articulo_ID", "articulo_id", "Articulo_Id") or "").strip()
                if self._is_total_row(code) or (not code and not articulo_id):
                    skipped += 1
                    continue

                rows.append(
                    VentaMensualRaw(
                        raw_id=str(uuid4()),
                        lote_id="",
                        fuente="ireks",
                        cliente_id=cliente_id,
                        periodo=periodo,
                        articulo_codigo_origen=code,
                        articulo_id=articulo_id,
                        articulo_descripcion_origen=str(
                            self._get_any(item, "Descripción", "DescripciÃ³n", "Descripcion", "descripcion") or ""
                        ).strip(),
                        venta_kilos=self._to_float(self._get_any(item, "venta_Kilos", "venta_kilos", "Kilos", "kilos")),
                        venta_kilos_sc=self._to_float(
                            self._get_any(item, "venta_kilos_SC", "venta_Kilos_SC", "S/C", "SC", "sc")
                        ),
                        venta_euros=self._to_float(
                            self._get_any(item, "Venta_euros", "venta_Euros", "venta_euros", "Ventas", "ventas")
                        ),
                        payload_json=json.dumps(item, ensure_ascii=False),
                    )
                )

            if not rows:
                return SalesOpResult(False, "No se pudo importar ninguna fila válida.", imported=0, incidencias=skipped)

            lote = VentaImportLote(
                lote_id=str(uuid4()),
                fuente="ireks",
                cliente_id=first_cliente_id,
                periodo=first_period,
                archivo_nombre=file_path.name,
                archivo_hash=file_hash,
                estado="procesado",
            )
            session.add(lote)
            session.flush()
            for row in rows:
                row.lote_id = lote.lote_id
                session.add(row)
            session.commit()

        return SalesOpResult(True, "Importación IREKS completada.", imported=len(rows), incidencias=skipped)

    def import_igsa_excel(self, file_path: Path) -> SalesOpResult:
        try:
            data = self._read_igsa_consolidado_rows(file_path)
        except ValueError as exc:
            return SalesOpResult(False, str(exc))
        if not data:
            return SalesOpResult(False, "El Excel IGSA no contiene filas válidas en hoja 'consolidado'.")

        file_hash = self._file_hash(file_path)
        with Session(engine) as session:
            existing = session.exec(
                select(VentaImportLote).where(
                    VentaImportLote.fuente == "igsa",
                    VentaImportLote.archivo_hash == file_hash,
                )
            ).first()
            was_reimport = existing is not None

            rows: list[VentaMensualRaw] = []
            skipped = 0
            first_period = ""
            first_cliente_id = ""
            client_rows = list(
                session.exec(
                    select(Cliente.cliente_id, Cliente.cliente_nombre_comercial, Cliente.cliente_nombre_fiscal, Cliente.cliente_tipo)
                )
            )
            allowed_by_id: dict[str, str] = {}
            allowed_by_name: dict[str, str] = {}
            for cid, ncom, nfis, ctype in client_rows:
                tipo = str(ctype or "").strip().lower()
                if tipo not in SALES_CLIENT_TYPES:
                    continue
                cliente_id_val = str(cid or "").strip()
                if not cliente_id_val:
                    continue
                allowed_by_id[cliente_id_val] = cliente_id_val
                for raw_name in (ncom, nfis):
                    norm_name = self._normalize_search_text(raw_name)
                    if norm_name:
                        allowed_by_name[norm_name] = cliente_id_val
            for item in data:
                year = self._to_int(item.get("anio"))
                month = self._to_int(item.get("mes_numero"))
                if year <= 0 or month < 1 or month > 12:
                    skipped += 1
                    continue
                periodo = f"{year:04d}-{month:02d}"
                if not first_period:
                    first_period = periodo
                cliente_id = self._resolve_igsa_cliente_id(
                    item.get("distribuidor_uuid"),
                    item.get("distribuidor_nombre"),
                    allowed_by_id,
                    allowed_by_name,
                )
                if cliente_id and not first_cliente_id:
                    first_cliente_id = cliente_id

                code = self._normalize_code(item.get("articulo_referencia"))
                articulo_id = str(item.get("articulo_id") or "").strip()
                if self._is_total_row(code) or (not code and not articulo_id):
                    skipped += 1
                    continue

                tipo = self._normalize_igsa_tipo(item.get("tipo"))
                kilos = self._to_float(item.get("kilos"))
                if tipo == "venta":
                    venta_kilos = kilos
                    venta_kilos_sc = 0.0
                elif tipo in {"s/c", "muestras"}:
                    venta_kilos = 0.0
                    venta_kilos_sc = kilos
                else:
                    skipped += 1
                    continue

                rows.append(
                    VentaMensualRaw(
                        raw_id=str(uuid4()),
                        lote_id="",
                        fuente="igsa",
                        cliente_id=cliente_id,
                        periodo=periodo,
                        articulo_codigo_origen=code,
                        articulo_id=articulo_id,
                        articulo_descripcion_origen=str(item.get("articulo_descripcion") or "").strip(),
                        venta_kilos=venta_kilos,
                        venta_kilos_sc=venta_kilos_sc,
                        venta_euros=0.0,
                        payload_json=json.dumps(item, ensure_ascii=False),
                    )
                )

            if not rows:
                return SalesOpResult(False, "No se pudo importar ninguna fila válida.", imported=0, incidencias=skipped)

            # Reemplaza periodos ya importados de IGSA para evitar duplicados
            # cuando se vuelven a importar meses/años (mismo cliente y periodo).
            period_set = {str(r.periodo or "").strip() for r in rows if str(r.periodo or "").strip()}
            for periodo in period_set:
                stmt = select(VentaMensualRaw).where(
                    col(VentaMensualRaw.fuente) == "igsa",
                    col(VentaMensualRaw.periodo) == periodo,
                )
                for old_row in session.exec(stmt):
                    session.delete(old_row)
            if was_reimport:
                old_lotes = list(
                    session.exec(
                        select(VentaImportLote).where(
                            VentaImportLote.fuente == "igsa",
                            VentaImportLote.archivo_hash == file_hash,
                        )
                    )
                )
                for old_lote in old_lotes:
                    session.delete(old_lote)

            lote = VentaImportLote(
                lote_id=str(uuid4()),
                fuente="igsa",
                cliente_id=first_cliente_id,
                periodo=first_period,
                archivo_nombre=file_path.name,
                archivo_hash=file_hash,
                estado="procesado",
            )
            session.add(lote)
            session.flush()
            for row in rows:
                row.lote_id = lote.lote_id
                session.add(row)
            session.commit()
            self._sync_igsa_sales_to_warehouse(session, rows)
            session.commit()

        self._igsa_rows_cache = None
        self._igsa_mtime_ns = None
        msg = "Importación IGSA completada."
        if was_reimport:
            msg = "Reimportación IGSA completada (periodos reemplazados)."
        return SalesOpResult(True, msg, imported=len(rows), incidencias=skipped)

    def parse_igsa_workbook_by_sheets(self, file_path: Path) -> tuple[list[IgsaWorkbookParsedLine], list[str]]:
        return self._igsa_workbook_flow_service.parse_igsa_workbook_by_sheets(file_path)

    def build_igsa_workbook_preview(
        self,
        lines: list[IgsaWorkbookParsedLine],
        cliente_id: str,
    ) -> tuple[list[dict[str, object]], list[str]]:
        return self._igsa_workbook_flow_service.build_igsa_workbook_preview(lines, cliente_id)

    def import_igsa_workbook_lines(
        self,
        lines: list[IgsaWorkbookParsedLine],
        cliente_id: str,
        *,
        force_reimport: bool = False,
    ) -> SalesOpResult:
        return self._igsa_workbook_flow_service.import_igsa_workbook_lines(
            lines,
            cliente_id,
            force_reimport=force_reimport,
            sync_warehouse_callback=self._sync_igsa_sales_to_warehouse,
        )

    def parse_igsa_pdf_files(self, file_paths: list[Path]) -> tuple[list[IgsaPdfParsedLine], list[str]]:
        return self._igsa_pdf_flow_service.parse_igsa_pdf_files(file_paths)

    def import_igsa_pdf_lines(self, lines: list[IgsaPdfParsedLine], cliente_id: str = "") -> SalesOpResult:
        return self._igsa_pdf_flow_service.import_igsa_pdf_lines(
            lines,
            cliente_id,
            sync_warehouse_callback=self._sync_igsa_sales_to_warehouse,
        )

    def rebuild_igsa_warehouse_movements(self, periodo: str = "") -> SalesOpResult:
        clean_periodo = str(periodo or "").strip()
        if clean_periodo and not re.fullmatch(r"\d{4}-\d{2}", clean_periodo):
            return SalesOpResult(False, "Periodo inválido. Usa formato AAAA-MM o vacío para todos.")

        with Session(engine) as session:
            stmt = select(VentaMensualRaw).where(col(VentaMensualRaw.fuente).in_(["igsa", "igsa_pdf"]))
            if clean_periodo:
                stmt = stmt.where(col(VentaMensualRaw.periodo) == clean_periodo)
            rows = list(session.exec(stmt))
            if not rows:
                return SalesOpResult(True, "No hay filas IGSA para regenerar.", imported=0, incidencias=0)
            self._sync_igsa_sales_to_warehouse(session, rows)
            session.commit()
        return SalesOpResult(
            True,
            f"Regeneración IGSA completada{' para ' + clean_periodo if clean_periodo else ''}.",
            imported=len(rows),
            incidencias=0,
        )

    def list_years(self) -> list[int]:
        return self._sales_annual_comparison_service.list_years()

    def list_years_igsa(self) -> list[int]:
        return self._sales_annual_comparison_service.list_years_igsa()

    def list_filter_clients(self) -> list[Cliente]:
        return self._sales_annual_comparison_service.list_filter_clients()

    def list_filter_products(self) -> list[IngredienteIreks]:
        return self._sales_annual_comparison_service.list_filter_products()

    def list_filter_manufacturers(self) -> list[Fabricante]:
        return self._sales_annual_comparison_service.list_filter_manufacturers()

    def list_filter_manufacturers_igsa(self) -> list[Fabricante]:
        return self._sales_annual_comparison_service.list_filter_manufacturers_igsa()

    def list_filter_families(self, fabricante_id: str = "") -> list[Familia]:
        return self._sales_annual_comparison_service.list_filter_families(fabricante_id)

    def list_filter_families_igsa(self, fabricante_id: str = "") -> list[Familia]:
        return self._sales_annual_comparison_service.list_filter_families_igsa(fabricante_id)

    def list_filter_subfamilies(self, familia_id: str = "") -> list[Subfamilia]:
        return self._sales_annual_comparison_service.list_filter_subfamilies(familia_id)

    def list_filter_subfamilies_igsa(self, familia_id: str = "") -> list[Subfamilia]:
        return self._sales_annual_comparison_service.list_filter_subfamilies_igsa(familia_id)

    def listar_resumen_anual_igsa(
        self,
        year: int,
        month: int = 0,
        acumulado: bool = False,
        producto_texto: str = "",
        fabricante_id: str = "",
        familia_id: str = "",
        subfamilia_id: str = "",
    ) -> list[SalesComparisonRow]:
        return self._sales_annual_comparison_service.listar_resumen_anual_igsa(
            year=year,
            month=month,
            acumulado=acumulado,
            producto_texto=producto_texto,
            fabricante_id=fabricante_id,
            familia_id=familia_id,
            subfamilia_id=subfamilia_id,
        )

    def listar_resumen_anual(
        self,
        year: int,
        month: int = 0,
        acumulado: bool = False,
        cliente_id: str = "",
        articulo_id: str = "",
        producto_texto: str = "",
        fabricante_id: str = "",
        familia_id: str = "",
        subfamilia_id: str = "",
    ) -> list[SalesComparisonRow]:
        return self._sales_annual_comparison_service.listar_resumen_anual(
            year=year,
            month=month,
            acumulado=acumulado,
            cliente_id=cliente_id,
            articulo_id=articulo_id,
            producto_texto=producto_texto,
            fabricante_id=fabricante_id,
            familia_id=familia_id,
            subfamilia_id=subfamilia_id,
        )

    def _read_json(self, file_path: Path):
        try:
            raw = self._decode_text(file_path.read_bytes())
            return json.loads(raw)
        except UnicodeDecodeError as exc:
            raise ValueError("No se pudo leer el JSON con una codificacion soportada.") from exc
        except json.JSONDecodeError as exc:
            raise ValueError(f"El archivo seleccionado no contiene JSON valido: linea {exc.lineno}, columna {exc.colno}.") from exc

    def _decode_text(self, content: bytes) -> str:
        if content.startswith((b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
            return content.decode("utf-32")
        if content.startswith((b"\xff\xfe", b"\xfe\xff")):
            return content.decode("utf-16")
        if content.startswith(b"\xef\xbb\xbf"):
            return content.decode("utf-8-sig")

        sample = content[:200]
        even_nulls = sample[0::2].count(0)
        odd_nulls = sample[1::2].count(0)
        if odd_nulls > even_nulls and odd_nulls >= max(2, len(sample) // 6):
            return content.decode("utf-16-le")
        if even_nulls > odd_nulls and even_nulls >= max(2, len(sample) // 6):
            return content.decode("utf-16-be")

        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8-sig")

    def _file_hash(self, file_path: Path) -> str:
        digest = hashlib.sha256()
        with file_path.open("rb") as fh:
            while True:
                chunk = fh.read(1024 * 1024)
                if not chunk:
                    break
                digest.update(chunk)
        return digest.hexdigest()

    def _get_any(self, data: dict, *keys: str):
        for key in keys:
            if key in data:
                return data[key]
        normalized = {self._normalize_key(k): v for k, v in data.items()}
        for key in keys:
            norm = self._normalize_key(key)
            if norm in normalized:
                return normalized[norm]
        return None

    def _normalize_key(self, value) -> str:
        text = str(value or "").strip().lower()
        normalized = unicodedata.normalize("NFD", text)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        return re.sub(r"[^a-z0-9]+", "", normalized)

    def _normalize_search_text(self, value) -> str:
        text = str(value or "").strip().lower()
        normalized = unicodedata.normalize("NFD", text)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        return re.sub(r"\s+", " ", normalized)

    def _normalize_code(self, value) -> str:
        text = str(value or "").strip().upper()
        if not text:
            return ""
        if re.fullmatch(r"\d+(\.0+)?", text):
            return str(int(float(text)))
        return text

    def _is_total_row(self, code: str) -> bool:
        return self._normalize_code(code) == "TOTAL"

    def _to_int(self, value) -> int:
        try:
            txt = str(value or "").strip()
            if not txt:
                return 0
            return int(float(txt.replace(",", ".")))
        except Exception:
            return 0

    def _to_float(self, value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            try:
                if isinstance(value, float) and math.isnan(value):
                    return 0.0
                return float(value)
            except Exception:
                return 0.0
        text = str(value).strip()
        if not text:
            return 0.0
        match = re.search(r"[-+]?\d+(?:[.,]\d+)?", text)
        if match is None:
            return 0.0
        try:
            return float(match.group(0).replace(",", "."))
        except Exception:
            return 0.0

    def _parse_month(self, value) -> int:
        text = str(value or "").strip().lower()
        if not text:
            return 0
        normalized = unicodedata.normalize("NFD", text)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        normalized = normalized.replace(".", "")
        month_map = {
            "enero": 1,
            "ene": 1,
            "febrero": 2,
            "feb": 2,
            "marzo": 3,
            "mar": 3,
            "abril": 4,
            "abr": 4,
            "mayo": 5,
            "may": 5,
            "junio": 6,
            "jun": 6,
            "julio": 7,
            "jul": 7,
            "agosto": 8,
            "ago": 8,
            "septiembre": 9,
            "setiembre": 9,
            "sep": 9,
            "set": 9,
            "octubre": 10,
            "oct": 10,
            "noviembre": 11,
            "nov": 11,
            "diciembre": 12,
            "dic": 12,
        }
        if normalized in month_map:
            return month_map[normalized]
        return self._to_int(normalized)

    def _normalize_igsa_tipo(self, value) -> str:
        raw = str(value or "").strip().lower()
        if not raw:
            return ""
        normalized = unicodedata.normalize("NFD", raw)
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        normalized = normalized.replace(" ", "")
        if normalized in {"venta", "ventas"}:
            return "venta"
        if normalized in {"s/c", "c/s", "sc"}:
            return "s/c"
        if normalized in {"muestra", "muestras"}:
            return "muestras"
        if normalized in {"promocion", "promociones"}:
            return "promociones"
        return normalized

    def _resolve_igsa_cliente_id(
        self,
        distribuidor_uuid: object,
        distribuidor_nombre: object,
        allowed_by_id: dict[str, str],
        allowed_by_name: dict[str, str],
    ) -> str:
        raw_id = str(distribuidor_uuid or "").strip()
        if raw_id and raw_id in allowed_by_id:
            return raw_id
        raw_name = self._normalize_search_text(distribuidor_nombre)
        if raw_name and raw_name in allowed_by_name:
            return allowed_by_name[raw_name]
        return raw_id

    def _normalize_ireks_json_payload(self, data: object) -> tuple[list[dict[str, object]], int, int, str]:
        default_year = 0
        default_month = 0
        default_cliente_id = ""
        rows_source: list[object] | None = None

        if isinstance(data, list):
            rows_source = data
        elif isinstance(data, dict):
            for key in ("articulos", "items", "filas", "rows", "lineas"):
                value = data.get(key)
                if isinstance(value, list):
                    rows_source = value
                    break
            cliente = data.get("cliente")
            if isinstance(cliente, dict):
                default_cliente_id = str(self._get_any(cliente, "id", "cliente_id", "Cliente_ID", "ClienteId") or "").strip()
            if not default_cliente_id:
                default_cliente_id = str(
                    self._get_any(data, "cliente_id", "Cliente_ID", "ClienteId", "distribuidor_uuid") or ""
                ).strip()
            periodo = data.get("periodo")
            if isinstance(periodo, dict):
                default_year = self._to_int(self._get_any(periodo, "anio", "año", "year"))
                default_month = self._parse_month(self._get_any(periodo, "mes", "month"))
                if default_month < 1:
                    default_month = self._to_int(self._get_any(periodo, "mes_numero", "numero_mes"))
            elif isinstance(periodo, str):
                raw_periodo = str(periodo or "").strip()
                match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", raw_periodo)
                if match is not None:
                    default_year = self._to_int(match.group(1))
                    default_month = self._to_int(match.group(2))

        if rows_source is None:
            return [], default_year, default_month, default_cliente_id
        return [item for item in rows_source if isinstance(item, dict)], default_year, default_month, default_cliente_id

    def _sync_igsa_sales_to_warehouse(self, session: Session, raw_rows: list[VentaMensualRaw]) -> None:
        article_ids = sorted(
            {
                str(getattr(r, "articulo_id", "") or "").strip()
                for r in raw_rows
                if str(getattr(r, "articulo_id", "") or "").strip()
            }
        )
        product_weight_by_id: dict[str, float] = {}
        if article_ids:
            products = list(
                session.exec(
                    select(IngredienteIreks).where(col(IngredienteIreks.articulo_id).in_(article_ids))
                )
            )
            for product in products:
                aid = str(getattr(product, "articulo_id", "") or "").strip()
                if not aid:
                    continue
                weight_total = float(getattr(product, "articulo_envase_peso_total", 0.0) or 0.0)
                weight_unit = float(getattr(product, "articulo_envase_peso", 0.0) or 0.0)
                product_weight_by_id[aid] = weight_total if weight_total > 0 else weight_unit

        # Caducidad por articulo+lote desde entradas de almacen ya registradas.
        expiry_by_key: dict[tuple[str, str, str], date | None] = {}
        if article_ids:
            entry_rows = list(
                session.exec(
                    select(AlmacenMovimiento).where(
                        col(AlmacenMovimiento.articulo_id).in_(article_ids),
                        col(AlmacenMovimiento.cantidad) > 0,
                    )
                )
            )
            for mov in entry_rows:
                almacen_id = str(getattr(mov, "almacen_id", "") or "").strip()
                articulo_id = str(getattr(mov, "articulo_id", "") or "").strip()
                lote = str(getattr(mov, "articulo_lote", "") or "").strip().upper()
                cad = getattr(mov, "articulo_caducidad", None)
                if not almacen_id or not articulo_id or not lote or cad is None:
                    continue
                key = (almacen_id, articulo_id, lote)
                prev = expiry_by_key.get(key)
                if prev is None or cad > prev:
                    expiry_by_key[key] = cad

        period_set = {str(r.periodo or "").strip() for r in raw_rows if str(r.periodo or "").strip()}
        for periodo in period_set:
            pedido_numero = f"IGSA-{periodo}"
            delete_stmt = select(AlmacenMovimiento).where(
                col(AlmacenMovimiento.pedido_numero) == pedido_numero
            )
            for mov in session.exec(delete_stmt):
                if str(getattr(mov, "albaran_item_id", "") or "").startswith("igsa:"):
                    session.delete(mov)

        for row in raw_rows:
            periodo = str(row.periodo or "").strip()
            almacen_id = str(row.cliente_id or "").strip()
            articulo_id = str(row.articulo_id or "").strip()
            if not periodo or not almacen_id or not articulo_id:
                continue

            payload = self._safe_json_dict(row.payload_json)
            cantidad_unidades = self._to_float(payload.get("cantidad_documento"))
            kilos_total = float(row.venta_kilos or 0.0) + float(row.venta_kilos_sc or 0.0)
            if cantidad_unidades <= 0 and kilos_total > 0:
                envase_peso = float(product_weight_by_id.get(articulo_id, 0.0) or 0.0)
                if envase_peso <= 0:
                    envase_peso = self._to_float(payload.get("envase_peso_documento"))
                if envase_peso <= 0:
                    envase_peso = self._to_float(payload.get("envase_peso"))
                if envase_peso > 0:
                    cantidad_unidades = kilos_total / envase_peso
            if cantidad_unidades <= 0:
                cantidad_unidades = self._to_float(payload.get("cantidad"))
            if cantidad_unidades <= 0:
                # Fallback: usa kilos cuando no hay cantidad.
                cantidad_unidades = kilos_total
            if cantidad_unidades <= 0:
                continue

            tipo = self._normalize_igsa_tipo(payload.get("tipo"))
            lote = str(payload.get("lote") or "").strip()
            lote_key = lote.strip().upper()
            caducidad = expiry_by_key.get((almacen_id, articulo_id, lote_key)) if lote_key else None
            mes = self._to_int(periodo.split("-")[1] if "-" in periodo else 0)
            anio = self._to_int(periodo.split("-")[0] if "-" in periodo else 0)
            fecha = date(anio if anio > 0 else date.today().year, mes if 1 <= mes <= 12 else 1, 1)
            unique_key = f"igsa:{periodo}:{almacen_id}:{articulo_id}:{lote}:{tipo}:{cantidad_unidades}"
            ref_id = str(uuid5(NAMESPACE_URL, unique_key))

            session.add(
                AlmacenMovimiento(
                    almacen_id=almacen_id,
                    articulo_id=articulo_id,
                    pedido_numero=f"IGSA-{periodo}",
                    pedido_albaran_numero=f"IGSA-{tipo}" if tipo else "IGSA",
                    cantidad=-abs(cantidad_unidades),
                    articulo_lote=lote,
                    articulo_caducidad=caducidad,
                    fecha_pedido=fecha,
                    albaran_item_id=f"igsa:{ref_id}",
                )
            )

    def _safe_json_dict(self, text: str) -> dict:
        raw = str(text or "").strip()
        if not raw:
            return {}
        try:
            data = json.loads(raw)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _igsa_file_path(self) -> Path:
        return BASE_DIR.parent / "Recursos" / "Ventas" / "IGSA - Ventas 01 Enero - 2026.xlsx"

    def _read_igsa_consolidado_rows(self, file_path: Path) -> list[dict[str, object]]:
        if not file_path.exists():
            return []
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style, apply openpyxl's default",
                category=UserWarning,
            )
            workbook = load_workbook(file_path, data_only=True)
        if "consolidado" not in workbook.sheetnames:
            return []
        sheet = workbook["consolidado"]
        ws = sheet if isinstance(sheet, Worksheet) else None
        if ws is None:
            return []
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_cells]
        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            item = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers) and headers[idx]}
            cantidad = self._to_float(item.get("cantidad"))
            envase_peso = self._to_float(item.get("envase_peso"))
            item["kilos"] = cantidad * envase_peso
            item["anio"] = self._to_int(item.get("anio"))
            item["mes_numero"] = self._to_int(item.get("mes_numero"))
            rows.append(item)
        return rows

    def _load_igsa_rows(self) -> list[dict[str, object]]:
        file_path = self._igsa_file_path()
        if not file_path.exists():
            return []
        stat = file_path.stat()
        mtime_ns = int(getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1_000_000_000)))
        if self._igsa_rows_cache is not None and self._igsa_mtime_ns == mtime_ns:
            return self._igsa_rows_cache
        rows = self._read_igsa_consolidado_rows(file_path)
        self._igsa_rows_cache = rows
        self._igsa_mtime_ns = mtime_ns
        return rows

