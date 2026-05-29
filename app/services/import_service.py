from pathlib import Path
import csv
import json
import unicodedata
import warnings
from typing import Any, Callable, cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class ImportService:
    def read_rows(self, file_path: Path) -> list[dict[str, Any]]:
        suffix = file_path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_excel(file_path)
        if suffix == ".csv":
            return self._read_csv(file_path)
        if suffix == ".json":
            return self._read_json(file_path)
        raise ValueError(f"Formato no soportado: {suffix}")

    def import_with_schema(
        self,
        file_path: Path,
        schema: list[dict[str, Any]],
        create_fn: Callable[[dict[str, Any]], None],
        required_fields: list[str] | None = None,
        aliases: dict[str, list[str]] | None = None,
    ) -> tuple[int, list[str]]:
        rows = self.read_rows(file_path)
        required = set(required_fields or [])
        aliases = aliases or {}
        imported = 0
        errors: list[str] = []

        for idx, row in enumerate(rows, start=2):
            try:
                payload = self.build_payload(row=row, schema=schema, aliases=aliases)
                self.validate_required_fields(payload, list(required))
                create_fn(payload)
                imported += 1
            except Exception as exc:
                errors.append(f"Fila {idx}: {exc}")
        return imported, errors

    def map_rows(
        self,
        file_path: Path,
        schema: list[dict[str, Any]],
        aliases: dict[str, list[str]] | None = None,
    ) -> list[dict[str, Any]]:
        rows = self.read_rows(file_path)
        return [self.build_payload(row=row, schema=schema, aliases=aliases or {}) for row in rows]

    def build_payload(
        self,
        row: dict[str, Any],
        schema: list[dict[str, Any]],
        aliases: dict[str, list[str]] | None = None,
    ) -> dict[str, Any]:
        return self._build_payload(row, schema, aliases or {})

    def validate_required_fields(self, payload: dict[str, Any], required_fields: list[str]) -> None:
        for req in required_fields:
            if not payload.get(req):
                raise ValueError(f"Campo obligatorio vacio: {req}")

    def _build_payload(
        self, row: dict[str, Any], schema: list[dict[str, Any]], aliases: dict[str, list[str]]
    ) -> dict[str, Any]:
        normalized_row = {self._normalize_key(k): v for k, v in row.items()}
        payload: dict[str, Any] = {}

        for field in schema:
            name = field["name"]
            candidates = [name, field.get("label", ""), *aliases.get(name, [])]
            raw_value = None
            for candidate in candidates:
                key = self._normalize_key(candidate)
                if key in normalized_row and normalized_row[key] not in (None, ""):
                    raw_value = normalized_row[key]
                    break
            payload[name] = self._coerce_value(raw_value, field)
        return payload

    def _coerce_value(self, raw_value: Any, field: dict[str, Any]) -> Any:
        ftype = field.get("type", "text")
        default = field.get("default")
        if raw_value in (None, ""):
            if default is not None:
                return default
            if ftype == "float":
                return 0.0
            if ftype == "bool":
                return False
            return ""

        if ftype == "float":
            value = str(raw_value).replace(",", ".").strip()
            return float(value)
        if ftype == "bool":
            value = str(raw_value).strip().lower()
            return value in {"1", "true", "si", "sí", "yes", "x"}
        return str(raw_value).strip()

    def _read_excel(self, file_path: Path) -> list[dict[str, Any]]:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style, apply openpyxl's default",
                category=UserWarning,
            )
            workbook = load_workbook(file_path, data_only=True)
        sheet = cast(Worksheet, workbook.active)
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_cells]
        rows: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            rows.append({headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)})
        return rows

    def _read_csv(self, file_path: Path) -> list[dict[str, Any]]:
        encodings = ["utf-8-sig", "latin-1"]
        for encoding in encodings:
            try:
                with file_path.open("r", encoding=encoding, newline="") as fh:
                    sample = fh.read(8192)
                    fh.seek(0)
                    delimiter = ","
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                        delimiter = str(getattr(dialect, "delimiter", ",") or ",")
                    except Exception:
                        delimiter = ";" if ";" in sample and "," not in sample.splitlines()[0] else ","
                    reader = csv.DictReader(fh, delimiter=delimiter)
                    return [row for row in reader if any((v or "").strip() for v in row.values())]
            except UnicodeDecodeError:
                continue
        raise ValueError("No se pudo leer el CSV con codificacion soportada.")

    def _read_json(self, file_path: Path) -> list[dict[str, Any]]:
        encodings = ["utf-8-sig", "latin-1"]
        data: Any = None
        for encoding in encodings:
            try:
                with file_path.open("r", encoding=encoding) as fh:
                    data = json.load(fh)
                break
            except UnicodeDecodeError:
                continue
        if data is None:
            raise ValueError("No se pudo leer el JSON con codificacion soportada.")
        if isinstance(data, list):
            rows = [row for row in data if isinstance(row, dict)]
            if not rows:
                raise ValueError("El JSON no contiene filas validas.")
            return rows
        if isinstance(data, dict):
            for key in ("rows", "items", "data", "albaranes_items"):
                candidate = data.get(key)
                if isinstance(candidate, list):
                    rows = [row for row in candidate if isinstance(row, dict)]
                    if rows:
                        return rows
            return [data]
        raise ValueError("Estructura JSON no soportada.")

    def _normalize_key(self, value: str) -> str:
        text = unicodedata.normalize("NFD", str(value))
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        return text.strip().lower().replace(" ", "_")
