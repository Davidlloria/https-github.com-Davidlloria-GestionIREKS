from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _empty_export_row() -> dict[str, str]:
    return {
        "articulo_id": "",
        "ref": "",
        "nombre": "",
        "lote": "",
        "caducidad": "",
        "teorico_uds": "0",
        "conteo_uds": "",
    }


@dataclass(slots=True)
class WarehouseCountTemplateFlowResult:
    status: str
    message: str = ""
    export_rows: list[dict[str, str]] = field(default_factory=list)
    imported_count: int = 0
    mapping: dict[tuple[str, str], str] = field(default_factory=dict)


class WarehouseCountTemplateFlowService:
    @staticmethod
    def count_template_column_indexes(header: list[str]) -> tuple[int, int, int]:
        idx_id = header.index("articulo_id") if "articulo_id" in header else -1
        idx_lote = header.index("lote") if "lote" in header else -1
        idx_conteo = header.index("conteo_uds") if "conteo_uds" in header else -1
        return idx_id, idx_lote, idx_conteo

    @staticmethod
    def count_template_mapping(
        rows: list[tuple[object, ...]],
        idx_id: int,
        idx_lote: int,
        idx_conteo: int,
    ) -> dict[tuple[str, str], str]:
        mapping: dict[tuple[str, str], str] = {}
        for row in rows:
            art_id = str(row[idx_id] or "").strip()
            lote = str(row[idx_lote] or "").strip()
            conteo_raw = str(row[idx_conteo] or "").strip()
            if not art_id or conteo_raw == "":
                continue
            mapping[(art_id, lote)] = conteo_raw
        return mapping

    @staticmethod
    def build_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
        export_rows: list[dict[str, str]] = []
        for row in rows:
            export_rows.append(
                {
                    "articulo_id": str(row.get("articulo_id") or ""),
                    "ref": str(row.get("ref") or ""),
                    "nombre": str(row.get("nombre") or ""),
                    "lote": str(row.get("lote") or ""),
                    "caducidad": str(row.get("caducidad") or ""),
                    "teorico_uds": str(row.get("teorico_uds") or "0"),
                    "conteo_uds": str(row.get("conteo_uds") or ""),
                }
            )
        return export_rows

    def read_import_file(self, file_path: Path) -> WarehouseCountTemplateFlowResult:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return WarehouseCountTemplateFlowResult(status="empty", message="El archivo no contiene datos.")
        header = [str(x or "").strip().lower() for x in rows[0]]
        idx_id, idx_lote, idx_conteo = self.count_template_column_indexes(header)
        if idx_id < 0 or idx_lote < 0 or idx_conteo < 0:
            return WarehouseCountTemplateFlowResult(
                status="missing_columns",
                message="Faltan columnas requeridas: articulo_id, lote, conteo_uds.",
            )
        mapping = self.count_template_mapping(rows[1:], idx_id, idx_lote, idx_conteo)
        if not mapping:
            return WarehouseCountTemplateFlowResult(status="empty", message="El archivo no contiene datos validos.")
        return WarehouseCountTemplateFlowResult(status="ready", imported_count=len(mapping), mapping=mapping)

    def prepare_export_rows(self, rows: list[dict[str, Any]]) -> WarehouseCountTemplateFlowResult:
        export_rows = self.build_export_rows(rows)
        if not export_rows:
            return WarehouseCountTemplateFlowResult(status="empty", message="No hay filas exportables.")
        return WarehouseCountTemplateFlowResult(status="ready", export_rows=export_rows)

    @staticmethod
    def normalize_lote(value: object) -> str:
        return str(value or "").strip()

    @staticmethod
    def normalize_caducidad(value: object) -> str:
        if isinstance(value, date):
            return value.isoformat()
        return str(value or "").strip()
