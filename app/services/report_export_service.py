from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.config import DATA_DIR


class ReportExportService:
    def default_path(self, title: str, suffix: str, folder: str = "listados_clientes") -> Path:
        reports_dir = DATA_DIR / "exports" / folder
        reports_dir.mkdir(parents=True, exist_ok=True)
        safe = "".join(ch if ch.isalnum() else "_" for ch in str(title or "listado").lower()).strip("_")
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return reports_dir / f"{safe[:40] or 'listado'}_{stamp}.{suffix.lstrip('.')}"

    def export_excel(self, path: str | Path, title: str, headers: list[str], rows: list[list[Any]], sheet_title: str = "Listado clientes") -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="3A78CF")
        for row in rows:
            ws.append(row)
        for col_idx in range(1, max(1, len(headers)) + 1):
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(2, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)
        ws.freeze_panes = "A3"
        wb.save(out)
        return out

    def export_pdf(self, path: str | Path, title: str, headers: list[str], rows: list[list[Any]]) -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(str(out), pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ListingTitleLeft",
            parent=styles["Title"],
            alignment=0,
        )
        story = [Paragraph(str(title or "Listado de clientes"), title_style), Spacer(1, 10)]
        table_data = [headers] + [[str(value) for value in row] for row in rows]
        column_count = max(1, len(headers))
        content_widths = [0] * column_count
        for row in table_data:
            for index in range(column_count):
                value = row[index] if index < len(row) else ""
                text = str(value or "")
                content_widths[index] = max(content_widths[index], len(text))
        min_widths = [24] * column_count
        max_widths = [max(40, min(220, width * 4 + 24)) for width in content_widths]
        available_width = doc.width
        scale = available_width / sum(max_widths)
        if scale < 1:
            col_widths = [max(min_widths[idx], width * scale) for idx, width in enumerate(max_widths)]
        else:
            col_widths = max_widths[:]
        width_delta = available_width - sum(col_widths)
        if width_delta != 0 and column_count:
            col_widths[-1] = max(min_widths[-1], col_widths[-1] + width_delta)
        table = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3A78CF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return out
