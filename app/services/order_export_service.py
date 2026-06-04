from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any, cast
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlmodel import Session, select

from app.core.config import BASE_DIR, PEDIDOS_EMAIL_DESTINO, PEDIDOS_HISTORICO_DIR
from app.core.database import engine
from app.models import Cliente, IngredienteIreks, Pedido, PedidoItem
from app.services.orders_mail_settings_service import OrdersMailSettingsService

@dataclass(frozen=True)
class OrderMailPreparation:
    attachment_path: Path
    preview: dict[str, str]


LEGAL_TEXT_PEDIDOS = (
    "PROTECCIÓN DE DATOS\n"
    "IREKS IBERICA, S.A. es el Responsable del tratamiento de los datos personales del Interesado y le informa que "
    "estos datos serán tratados de conformidad con lo dispuesto en las normativas vigentes en protección de datos "
    "personales, el Reglamento (UE) 2016/679 de 27 de abril de 2016 (GDPR). Dicho tratamiento se realizará para "
    "mantener una relación comercial mediante el envio de comunicaciones de nuestros productos o servicios y los datos "
    "se conservarán indefinidamente para fines de archivo mientras exista un interés mutuo para ello. No está previsto "
    "comunicar los datos a terceros (salvo obligación legal), y si fuera necesario hacerlo para la ejecución del "
    "contrato, se informará previamente al Interesado.\n"
    "Se informa al Interesado que tiene derecho a retirar el consentimiento para tratar los datos en cualquier momento "
    "y que, si ejerce este derecho, se deberá proceder a la rescisión del contrato en los términos expuestos en el "
    "mismo ya que el tratamiento de datos es imprescindible para la ejecución del contrato.\n"
    "De la misma forma también podrá ejercer los derechos de acceso, rectificación, supresión y portabilidad de sus "
    "datos y los de limitación u oposición al tratamiento y, si considera que el\n"
    "tratamiento de datos personales no se ajusta a la normativa vigente, también tiene derecho a presentar una "
    "reclamación ante la Autoridad de control (www.agpd.es)."
)


class OrderExportService:
    def __init__(self) -> None:
        self.orders_mail_settings = OrdersMailSettingsService()

    def build_order_workbook(self, pedido_id: str) -> tuple[Workbook, str]:
        with Session(engine) as session:
            pedido = session.get(Pedido, pedido_id)
            if pedido is None:
                raise ValueError("Pedido no encontrado.")
            cliente = session.get(Cliente, str(getattr(pedido, "almacen_id", "") or "").strip())
            rows = list(
                session.exec(
                    select(PedidoItem, IngredienteIreks)
                    .outerjoin(IngredienteIreks, cast(Any, IngredienteIreks.articulo_id == PedidoItem.articulo_id))
                    .where(PedidoItem.pedido_id == pedido_id)
                    .order_by(PedidoItem.item_id)
                )
            )
        pedido_numero = str(getattr(pedido, "pedido_numero", "") or "").strip()
        pedido_fecha = self.parse_date(getattr(pedido, "pedido_fecha", None))
        cliente_nombre = (
            str(getattr(cliente, "cliente_nombre_comercial", "") or "").strip()
            or str(getattr(cliente, "cliente_nombre_fiscal", "") or "").strip()
            or str(getattr(pedido, "almacen_id", "") or "").strip()
        )
        base_name = self.build_export_base_name(pedido_fecha, cliente)

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("No se pudo crear la hoja activa del libro.")
        ws.title = "PEDIDO"
        self.insert_order_logos(ws)
        ws["A1"] = "Formulario de pedidos IREKS v.1.24"
        ws["A1"].font = Font(name="Aptos Narrow", size=13, bold=True, color="FF71253E")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["C4"] = "Nº de pedido: "
        ws["D4"] = pedido_numero
        ws.merge_cells("D4:E4")
        ws["C5"] = "Cliente / Distribuidor: "
        ws["D5"] = cliente_nombre
        ws.merge_cells("D5:E5")
        ws["C6"] = "Fecha del pedido: "
        ws["D6"] = pedido_fecha
        ws["D6"].number_format = "mm-dd-yy"
        ws.merge_cells("D6:E6")
        ws["C7"] = "Fecha de entrega deseada: "
        ws["D7"] = None
        ws["D7"].number_format = "mm-dd-yy"
        ws.merge_cells("D7:E7")
        for cell_ref in ("C4", "C5", "C6", "C7", "D4", "D5", "D6", "D7"):
            ws[cell_ref].font = Font(name="Aptos Narrow", size=11, bold=False)
        for cell_ref in ("C4", "C5", "C6", "C7"):
            ws[cell_ref].alignment = Alignment(horizontal="right", vertical="center")
        for cell_ref in ("D4", "D5", "D6", "D7"):
            ws[cell_ref].alignment = Alignment(horizontal="center")

        ws["A9"] = "Cod. Artículo"
        ws["B9"] = "Producto"
        ws["C9"] = "kg /Und."
        ws["D9"] = "Nº Envases"
        ws["E9"] = "kg Total"

        header_fill = PatternFill("solid", fgColor="FFFDF6D8")
        for cell_ref in ("A9", "B9", "C9", "D9", "E9"):
            ws[cell_ref].fill = header_fill
            ws[cell_ref].font = Font(name="Aptos Narrow", size=11, color="FF71253E", bold=False)
            ws[cell_ref].alignment = Alignment(horizontal="center")
        ws["C9"].number_format = "#,##0.00_);[Red](#,##0.00)"
        ws["E9"].number_format = '_-* #,##0.00\\ "Kg."_-;\\-* #,##0.00\\ "Kg."_-;_-* "-"??\\ "Kg."_-;_-@_-'

        row_idx = 10
        total_kg = 0.0
        for item, article in rows:
            qty = float(getattr(item, "articulo_cantidad", 0.0) or 0.0)
            if qty <= 0:
                continue
            ref = (
                str(getattr(article, "articulo_referencia_corta", "") or "").strip()
                or str(getattr(article, "articulo_referencia", "") or "").strip()
                or str(getattr(item, "articulo_id", "") or "").strip()
            )
            nombre = str(getattr(article, "articulo_descripcion", "") or "").strip() if article else ref
            peso = float(getattr(article, "articulo_envase_peso_total", 0.0) or 0.0) if article else 0.0
            kg = qty * peso
            ws.cell(row=row_idx, column=1, value=ref)
            ws.cell(row=row_idx, column=2, value=nombre)
            ws.cell(row=row_idx, column=3, value=peso)
            ws.cell(row=row_idx, column=4, value=qty)
            ws.cell(row=row_idx, column=5, value=kg)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=3).number_format = "#,##0.00_);[Red](#,##0.00)"
            ws.cell(row=row_idx, column=4).number_format = "0"
            ws.cell(row=row_idx, column=5).number_format = '_-* #,##0.00\\ "Kg."_-;\\-* #,##0.00\\ "Kg."_-;_-* "-"??\\ "Kg."_-;_-@_-'
            for column_idx in range(1, 6):
                ws.cell(row=row_idx, column=column_idx).font = Font(name="Aptos Narrow", size=11, bold=False)
            total_kg += kg
            row_idx += 1

        data_last_row = row_idx - 1
        if data_last_row >= 10:
            table_name = f"PedidoItems_{''.join(ch for ch in str(pedido_id) if ch.isalnum())[:8] or 'X'}"
            table = Table(displayName=table_name, ref=f"A9:E{data_last_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            line_border = Border(
                left=Side(style="thin", color="FFD2B48C"),
                right=Side(style="thin", color="FFD2B48C"),
                top=Side(style="thin", color="FFD2B48C"),
                bottom=Side(style="thin", color="FFD2B48C"),
            )
            for row_num in range(9, data_last_row + 1):
                for column_idx in range(1, 6):
                    ws.cell(row_num, column_idx).border = line_border

        total_row = row_idx + 1
        total_fill = PatternFill("solid", fgColor="FFFDF6D8")
        ws.cell(row=total_row, column=4).fill = total_fill
        ws.cell(row=total_row, column=4, value="Total kg").font = Font(
            name="Aptos Narrow", size=11, bold=True, color="FF71253E"
        )
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=total_row, column=5, value=total_kg).number_format = '#,##0.00 "kg"'
        ws.cell(row=total_row, column=5).font = Font(name="Aptos Narrow", size=11, bold=True, color="FF71253E")
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        footer_row = row_idx + 3
        ws.merge_cells(f"A{footer_row}:E{footer_row}")
        footer_cell = ws.cell(row=footer_row, column=1, value=LEGAL_TEXT_PEDIDOS)
        footer_cell.font = Font(name="Aptos Narrow", size=4.5, bold=False)
        footer_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[footer_row].height = 62.25

        ws.column_dimensions["A"].width = 12.33
        ws.column_dimensions["B"].width = 32.07
        ws.column_dimensions["C"].width = 9.40
        ws.column_dimensions["D"].width = 11.47
        ws.column_dimensions["E"].width = 12.53
        ws.row_dimensions[1].height = 16.9
        ws.row_dimensions[9].height = 14.65
        return wb, base_name

    def build_export_base_name(self, pedido_fecha: date, cliente: Cliente | None) -> str:
        week = int(pedido_fecha.isocalendar().week)
        cliente_comercial = str(getattr(cliente, "cliente_nombre_comercial", "") or "").strip().lower()
        cliente_fiscal = str(getattr(cliente, "cliente_nombre_fiscal", "") or "").strip().lower()
        is_igsa = "igsa" in cliente_comercial or "igsa" in cliente_fiscal
        prefix = "LPA" if is_igsa else self.cliente_export_code(cliente)
        return f"{prefix} - Pedido SEMANA {week} - {pedido_fecha.strftime('%d.%m.%Y')}"

    @staticmethod
    def cliente_export_code(cliente: Cliente | None) -> str:
        if cliente is None:
            return "PED"
        abreviatura = str(getattr(cliente, "cliente_abreviatura", "") or "").strip().upper()
        if abreviatura:
            return abreviatura
        interno = str(getattr(cliente, "cliente_nombre_interno", "") or "").strip().upper()
        if interno:
            return interno
        name = (
            str(getattr(cliente, "cliente_nombre_comercial", "") or "").strip()
            or str(getattr(cliente, "cliente_nombre_fiscal", "") or "").strip()
        )
        words = [w for w in re.split(r"[^A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9]+", name) if w]
        if not words:
            return "PED"
        code = "".join(word[0].upper() for word in words if word)
        if len(code) < 3:
            tail = words[-1][1:].upper() if words[-1] else ""
            code = (code + tail)[:3]
        return (code or "PED")[:6]

    @staticmethod
    def insert_order_logos(ws: Any) -> None:
        assets_dir = BASE_DIR / "assets" / "logos" / "pedidos"
        for path, anchor in ((assets_dir / "Logo_cabecera_Pedidos_S.png", "A2"),):
            if not path.exists():
                continue
            try:
                img = XLImage(str(path))
                img.width = 254
                img.height = 129
                ws.add_image(img, anchor)
            except Exception:
                continue

    def save_order_excel_history(self, pedido_id: str, wb: Workbook, default_base_name: str) -> Path:
        with Session(engine) as session:
            pedido = session.get(Pedido, pedido_id)
            if pedido is None:
                raise ValueError("Pedido no encontrado para histórico.")
            pedido_fecha = self.parse_date(getattr(pedido, "pedido_fecha", None))
        base_name = self.sanitize_filename(default_base_name)
        folder = self.pedido_history_dir(pedido_fecha)
        folder.mkdir(parents=True, exist_ok=True)
        version = self.next_history_version(folder, base_name)
        target_path = folder / f"{base_name}_v{version}.xlsx"
        wb.save(target_path)
        return target_path

    def mark_order_exported(self, pedido_id: str) -> None:
        with Session(engine) as session:
            pedido = session.get(Pedido, pedido_id)
            if pedido is not None:
                pedido.pedido_estado = "E"
                session.add(pedido)
                session.commit()

    def build_order_mail_preview(self, pedido_id: str, pedido_numero: str, destino_email: str) -> dict[str, str]:
        with Session(engine) as session:
            pedido = session.get(Pedido, pedido_id)
            pedido_fecha = self.parse_date(getattr(pedido, "pedido_fecha", None)) if pedido is not None else date.today()
        numero = str(pedido_numero or "").strip() or "S/N"
        _ = numero
        semana = int(pedido_fecha.isocalendar().week)
        subject = f"Pedido Semana {semana}"
        body = (
            "Buenos días Vanessa\n"
            f"Adjunto el pedido Semana {semana} ({pedido_fecha.strftime('%d/%m/%Y')}).\n"
            "Saludos"
        )
        return {"to_email": destino_email, "subject": subject, "body": body}

    def prepare_order_mail_attachment(self, pedido_id: str, pedido_numero: str, destino_email: str) -> OrderMailPreparation:
        wb, default_base_name = self.build_order_workbook(pedido_id)
        attachment_path = self.save_order_excel_history(pedido_id, wb, default_base_name)
        preview = self.build_order_mail_preview(pedido_id, pedido_numero, destino_email)
        return OrderMailPreparation(attachment_path=attachment_path, preview=preview)

    def send_order_mail(
        self,
        *,
        pedido_id: str,
        pedido_numero: str,
        attachment_path: Path,
        destino_email: str,
        send_direct: bool,
        subject: str,
        body: str,
    ) -> dict[str, str]:
        try:
            outcome = self.open_outlook_mail_with_attachment(
                pedido_id=pedido_id,
                pedido_numero=pedido_numero,
                attachment_path=attachment_path,
                destino_email=destino_email,
                send_direct=send_direct,
                subject=subject,
                body=body,
            )
        except Exception as exc:  # noqa: BLE001
            self.log_order_mail_event(
                pedido_id=pedido_id,
                pedido_numero=pedido_numero,
                destino_email=destino_email,
                asunto="",
                adjunto_path=str(attachment_path),
                modo_envio="send" if send_direct else "draft",
                estado="ERROR",
                error_detalle=str(exc),
            )
            raise
        self.log_order_mail_event(
            pedido_id=pedido_id,
            pedido_numero=pedido_numero,
            destino_email=destino_email,
            asunto=str(outcome.get("subject") or "").strip(),
            adjunto_path=str(attachment_path),
            modo_envio="send" if send_direct else "draft",
            estado="ENVIADO" if send_direct else "BORRADOR",
            error_detalle="",
        )
        return outcome

    def open_outlook_mail_with_attachment(
        self,
        *,
        pedido_id: str,
        pedido_numero: str,
        attachment_path: Path,
        destino_email: str,
        send_direct: bool,
        subject: str,
        body: str,
    ) -> dict[str, str]:
        try:
            import win32com.client  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("No se encontró pywin32. Instala dependencias para habilitar Outlook.") from exc

        with Session(engine) as session:
            pedido = session.get(Pedido, pedido_id)
            pedido_fecha = self.parse_date(getattr(pedido, "pedido_fecha", None)) if pedido is not None else date.today()
            cliente = session.get(Cliente, str(getattr(pedido, "almacen_id", "") or "").strip()) if pedido is not None else None
            cliente_nombre = (
                str(getattr(cliente, "cliente_nombre_comercial", "") or "").strip()
                or str(getattr(cliente, "cliente_nombre_fiscal", "") or "").strip()
                or str(getattr(pedido, "almacen_id", "") or "").strip()
            )

        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = destino_email.strip() or PEDIDOS_EMAIL_DESTINO.strip()
        final_subject = str(subject or "").strip() or f"Pedido {str(pedido_numero or '').strip() or 'S/N'} - {cliente_nombre}"
        final_body = str(body or "").strip()
        if not final_body:
            final_body = (
                f"Hola,\n\n"
                f"Adjuntamos el pedido {str(pedido_numero or '').strip() or 'S/N'} ({pedido_fecha.strftime('%d/%m/%Y')}).\n\n"
                "Gracias."
            )
        mail.Subject = final_subject
        mail.Body = final_body
        mail.Attachments.Add(str(attachment_path))
        if send_direct:
            mail.Send()
        else:
            mail.Display()
        return {"subject": final_subject}

    def log_order_mail_event(
        self,
        *,
        pedido_id: str,
        pedido_numero: str,
        destino_email: str,
        asunto: str,
        adjunto_path: str,
        modo_envio: str,
        estado: str,
        error_detalle: str,
    ) -> None:
        try:
            with engine.begin() as conn:
                conn.exec_driver_sql(
                    """
                    INSERT INTO pedidos_email_log (
                        log_id, pedido_id, pedido_numero, destino_email, asunto, adjunto_path,
                        modo_envio, estado, error_detalle, creado_en
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        str(pedido_id or "").strip(),
                        str(pedido_numero or "").strip(),
                        str(destino_email or "").strip(),
                        str(asunto or "").strip(),
                        str(adjunto_path or "").strip(),
                        str(modo_envio or "").strip(),
                        str(estado or "").strip(),
                        str(error_detalle or "").strip(),
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )
        except Exception:
            return

    def pedido_history_dir(self, pedido_fecha: date) -> Path:
        loaded = self.orders_mail_settings.load()
        configured_base = str(loaded.get("historico_dir") or "").strip()
        base_dir = Path(configured_base) if configured_base else PEDIDOS_HISTORICO_DIR
        return base_dir / f"{int(pedido_fecha.year):04d}" / f"{int(pedido_fecha.month):02d}"

    @staticmethod
    def sanitize_filename(value: str) -> str:
        cleaned = re.sub(r'[<>:"/\\|?*]+', "_", str(value or "").strip())
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
        return cleaned or "pedido"

    @staticmethod
    def next_history_version(folder: Path, base_name: str) -> int:
        pattern = re.compile(rf"^{re.escape(base_name)}_v(\d+)\.xlsx$", re.IGNORECASE)
        max_version = 0
        for item in folder.glob(f"{base_name}_v*.xlsx"):
            match = pattern.match(item.name)
            if not match:
                continue
            try:
                version = int(match.group(1))
            except Exception:
                continue
            max_version = max(max_version, version)
        return max_version + 1

    @staticmethod
    def parse_date(value: object) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or "").strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except Exception:
                continue
        return date.today()
